"""Fiyat listesi Excel round-trip'i (spec §12A.1, §12A.2).

**Kural: export edilen dosya = import şablonu.** İkinci bir format yok. Bu yüzden
export ve import aynı sütun tanımından (`COLUMNS`) beslenir; sütun eklendiğinde iki
taraf birlikte değişir.

`Net Kâr` ve `Marj %` export'ta bilgi amaçlı yazılır, import'ta **yok sayılır** —
kârın tek doğruluk kaynağı motordur (CLAUDE.md §1).

Maliyet ve fiyat değişiklikleri versiyonlu yazılır (`sku_costs`, `sku_prices`);
geçmiş kayıt güncellenmez, `effective_from` ile yeni kayıt eklenir — böylece eski
siparişlerin kâr hesabı bozulmaz.
"""

from __future__ import annotations

import uuid
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.logging import get_logger
from app.engine.profit import LineInput, compute_line_profit
from app.models.catalog import Product, SkuCost, SkuLogistics, SkuPrice
from app.models.enums import CostSource, DraftStatus
from app.models.identity import Channel, Store
from app.models.workspace import ProductDraft
from app.services.commission import resolve_commission
from app.services.isolation import CROSS_BRAND_MESSAGE, belongs_to_another_brand

log = get_logger("services.pricelist")

ZERO = Decimal("0")

TEMPLATE_VERSION = "kavun-template-v1"
META_CELL = "A1"
HEADER_ROW = 2
FIRST_DATA_ROW = 3
SHEET_NAME = "Fiyat Listesi"
ERROR_SHEET_NAME = "Hatalar"

# Sütun sırası şablonun sözleşmesidir; değişirse `TEMPLATE_VERSION` de değişmelidir.
COLUMNS = (
    "SKU",
    "Ürün Adı",
    "Marka",
    "Kanal",
    "KDV %",
    "Desi",
    "Alış Maliyeti",
    "Satış Fiyatı",
    "Komisyon %",
    "Kargo (tahmini)",
    "Hizmet Bedeli",
    "Net Kâr",
    "Marj %",
)
# Import'ta okunmayan, yalnızca gösterim için yazılan sütunlar (spec §12A.1).
COMPUTED_COLUMNS = ("Net Kâr", "Marj %")

COLUMN_WIDTHS = (18, 40, 12, 14, 8, 8, 14, 14, 12, 16, 14, 14, 10)

# Türkiye'de geçerli KDV oranları; dışındaki değer satırı reddeder (spec §12A.2.3).
VALID_VAT_RATES = (Decimal("0"), Decimal("1"), Decimal("10"), Decimal("20"))


class TemplateError(ValueError):
    """Dosya Kavun şablonu değil ya da sürümü uyumsuz."""


@dataclass
class RowResult:
    """Bir satırın import sonucu — diff önizlemesinin kaynağı (spec §12A.2.1)."""

    row_no: int
    sku: str
    channel: str
    action: str  # yeni | guncelleme | degisiklik_yok | taslak | hata
    message: str = ""
    changes: dict[str, str] = field(default_factory=dict)


@dataclass
class ImportSummary:
    """Import özeti."""

    dry_run: bool
    yeni: int = 0
    guncelleme: int = 0
    degisiklik_yok: int = 0
    taslak: int = 0
    hata: int = 0
    rows: list[RowResult] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        """Log/JSON dostu özet."""
        return {
            "dry_run": self.dry_run,
            "yeni": self.yeni,
            "guncelleme": self.guncelleme,
            "degisiklik_yok": self.degisiklik_yok,
            "taslak": self.taslak,
            "hata": self.hata,
        }


@dataclass
class ParsedRow:
    """Dosyadan okunan ham satır (henüz doğrulanmadı)."""

    row_no: int
    values: dict[str, Any]


# --- ortak yardımcılar -------------------------------------------------------


def _decimal(value: Any, field_name: str) -> Decimal | None:
    """Hücreyi `Decimal`a çevirir; boşsa `None`, bozuksa hata fırlatır."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        # Excel'den float gelebilir; metin üzerinden geçmek kayan nokta hatasını önler.
        return Decimal(str(value).replace(",", ".").strip())  # allow-float: openpyxl hücresi
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{field_name}: sayı değil ({value!r})") from exc


def _text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _effective_cost(session: Session, product_id: uuid.UUID) -> SkuCost | None:
    return session.scalar(
        select(SkuCost)
        .where(SkuCost.product_id == product_id)
        .order_by(SkuCost.effective_from.desc(), SkuCost.created_at.desc())
        .limit(1)
    )


def _effective_price(
    session: Session, product_id: uuid.UUID, store_id: uuid.UUID
) -> SkuPrice | None:
    return session.scalar(
        select(SkuPrice)
        .where(SkuPrice.product_id == product_id, SkuPrice.store_id == store_id)
        .order_by(SkuPrice.effective_from.desc(), SkuPrice.created_at.desc())
        .limit(1)
    )


def _effective_logistics(session: Session, product_id: uuid.UUID) -> SkuLogistics | None:
    return session.scalar(
        select(SkuLogistics)
        .where(SkuLogistics.product_id == product_id)
        .order_by(SkuLogistics.effective_from.desc())
        .limit(1)
    )


def _stores_by_channel(session: Session) -> dict[str, Store]:
    """Aktif markanın kanal kodu → mağaza eşlemesi (guard markayı zaten kısıtlar)."""
    rows = session.execute(
        select(Store, Channel).join(Channel, Channel.id == Store.channel_id)
    ).all()
    return {channel.code.value: store for store, channel in rows}


def preview_profit(
    session: Session,
    product: Product,
    store: Store,
    *,
    price: Decimal | None,
    unit_cost: Decimal | None,
    cargo: Decimal,
    on_date: date,
) -> tuple[Decimal, Decimal, Decimal | None]:
    """Fiyat listesindeki bilgi sütunları: (net kâr, marj %, komisyon oranı).

    Motorun kendisi çağrılır — Excel'de ikinci bir kâr formülü yaşamaz.
    """
    if price is None or price <= ZERO:
        return ZERO, ZERO, None
    rate_row, source = resolve_commission(
        session, store_id=store.id, product=product, on_date=on_date
    )
    breakdown = compute_line_profit(
        LineInput(
            line_gross=price,
            qty=1,
            vat_percent=product.vat_rate,
            unit_cost_net=unit_cost,
            commission_rate=rate_row.rate if rate_row else None,
            commission_source=source,
            cargo_cost=cargo,
            service_fee=store.service_fee_per_order or ZERO,
        )
    )
    return breakdown.profit, breakdown.margin_pct, rate_row.rate if rate_row else None


@dataclass
class PriceRow:
    """Ürün çalışma alanı tablosunun bir satırı (ekran + export aynı kaynaktan)."""

    product_id: uuid.UUID
    sku: str
    name: str
    channel: str
    vat_rate: Decimal
    desi: Decimal | None
    unit_cost: Decimal | None
    price: Decimal | None
    commission_rate: Decimal | None
    service_fee: Decimal
    profit: Decimal
    margin_pct: Decimal


def price_rows(session: Session, *, today: date) -> list[PriceRow]:
    """Fiyat listesi satırları — ekran ve export bunu paylaşır (tek kaynak)."""
    rows: list[PriceRow] = []
    stores = _stores_by_channel(session)
    for product in session.scalars(select(Product).order_by(Product.sku)).all():
        cost = _effective_cost(session, product.id)
        logistics = _effective_logistics(session, product.id)
        for channel_code, store in sorted(stores.items()):
            price_row = _effective_price(session, product.id, store.id)
            price = price_row.price if price_row else None
            profit, margin, commission = preview_profit(
                session,
                product,
                store,
                price=price,
                unit_cost=cost.unit_cost if cost else None,
                cargo=ZERO,
                on_date=today,
            )
            rows.append(
                PriceRow(
                    product_id=product.id,
                    sku=product.sku,
                    name=product.name,
                    channel=channel_code,
                    vat_rate=product.vat_rate,
                    desi=logistics.desi if logistics else None,
                    unit_cost=cost.unit_cost if cost else None,
                    price=price,
                    commission_rate=commission,
                    service_fee=store.service_fee_per_order or ZERO,
                    profit=profit if price else ZERO,
                    margin_pct=margin if price else ZERO,
                )
            )
    return rows


# --- export ------------------------------------------------------------------


def export_price_list(session: Session, *, brand_name: str, today: date) -> bytes:
    """Aktif markanın fiyat listesini xlsx olarak üretir (spec §12A.1)."""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = SHEET_NAME

    # Şablon sürümü ilk satırda; import bunu kontrol eder.
    sheet[META_CELL] = TEMPLATE_VERSION
    sheet.row_dimensions[1].hidden = True

    for index, name in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=HEADER_ROW, column=index, value=name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[get_column_letter(index)].width = COLUMN_WIDTHS[index - 1]
    sheet.freeze_panes = f"A{FIRST_DATA_ROW}"

    for row_no, row in enumerate(price_rows(session, today=today), start=FIRST_DATA_ROW):
        values = (
            row.sku,
            row.name,
            brand_name,
            row.channel,
            row.vat_rate,
            row.desi,
            row.unit_cost,
            row.price,
            row.commission_rate * Decimal(100) if row.commission_rate is not None else None,
            ZERO,
            row.service_fee,
            row.profit if row.price else None,
            row.margin_pct if row.price else None,
        )
        for index, value in enumerate(values, start=1):
            sheet.cell(row=row_no, column=index, value=value)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# --- import ------------------------------------------------------------------


def parse_price_list(payload: bytes) -> list[ParsedRow]:
    """Dosyayı okur ve şablon sürümünü doğrular; hücreleri henüz yorumlamaz."""
    try:
        workbook = load_workbook(BytesIO(payload), data_only=True)
    except Exception as exc:  # openpyxl çeşitli hata tipleri fırlatır
        raise TemplateError("Dosya okunamadı; Excel (.xlsx) bekleniyor.") from exc

    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
    if sheet is None:
        raise TemplateError("Dosyada sayfa bulunamadı.")
    if _text(sheet[META_CELL].value) != TEMPLATE_VERSION:
        raise TemplateError(
            f"Şablon sürümü uyumsuz. Beklenen: {TEMPLATE_VERSION}. "
            "Güncel şablonu 'Excel'e Aktar' ile indirin."
        )

    headers = [_text(cell.value) for cell in sheet[HEADER_ROW]]
    missing = [name for name in COLUMNS if name not in headers]
    if missing:
        raise TemplateError(f"Eksik sütun(lar): {', '.join(missing)}")

    rows: list[ParsedRow] = []
    for row_no, row in enumerate(
        sheet.iter_rows(min_row=FIRST_DATA_ROW, values_only=True), start=FIRST_DATA_ROW
    ):
        if all(value is None or _text(value) == "" for value in row):
            continue
        values = {
            name: row[headers.index(name)] for name in COLUMNS if headers.index(name) < len(row)
        }
        rows.append(ParsedRow(row_no=row_no, values=values))
    return rows


def _validate(
    row: ParsedRow, stores: dict[str, Store], *, as_draft: bool = False
) -> tuple[dict[str, Any], str]:
    """Satırı doğrular; `(temiz değerler, hata mesajı)` döner.

    `as_draft=True` iken SKU'suz satırlar hata sayılmaz, taslak adayı olur (spec §12A.3).
    """
    sku = _text(row.values.get("SKU"))
    if not sku and not as_draft:
        return {}, "SKU boş"
    if not sku and not _text(row.values.get("Ürün Adı")):
        return {}, "SKU ve Ürün Adı birlikte boş olamaz"

    channel = _text(row.values.get("Kanal")).lower()
    if channel not in stores:
        known = ", ".join(sorted(stores)) or "-"
        return {}, f"Bilinmeyen kanal: {channel or '(boş)'} (tanımlı: {known})"

    try:
        vat = _decimal(row.values.get("KDV %"), "KDV %")
        desi = _decimal(row.values.get("Desi"), "Desi")
        cost = _decimal(row.values.get("Alış Maliyeti"), "Alış Maliyeti")
        price = _decimal(row.values.get("Satış Fiyatı"), "Satış Fiyatı")
    except ValueError as exc:
        return {}, str(exc)

    if vat is None:
        return {}, "KDV % zorunlu"
    if vat not in VALID_VAT_RATES:
        allowed = ", ".join(f"%{value}" for value in VALID_VAT_RATES)
        return {}, f"Geçersiz KDV oranı: %{vat} (geçerli: {allowed})"
    for label, value in (("Alış Maliyeti", cost), ("Satış Fiyatı", price), ("Desi", desi)):
        if value is not None and value < ZERO:
            return {}, f"{label} negatif olamaz"

    return {
        "sku": sku,
        "name": _text(row.values.get("Ürün Adı")) or sku,
        "is_draft": not sku,
        "channel": channel,
        "vat": vat,
        "desi": desi,
        "cost": cost,
        "price": price,
    }, ""


def _apply(
    session: Session,
    clean: dict[str, Any],
    store: Store,
    *,
    today: date,
    user: str | None,
    dry_run: bool,
) -> tuple[str, dict[str, str]]:
    """Bir satırı uygular; `(aksiyon, değişiklikler)` döner. `dry_run`'da yazma yok."""
    product = session.scalar(select(Product).where(Product.sku == clean["sku"]))
    changes: dict[str, str] = {}

    if product is None:
        changes["ürün"] = f"yeni: {clean['name']}"
        if clean["cost"] is not None:
            changes["maliyet"] = f"→ {clean['cost']}"
        if clean["price"] is not None:
            changes["fiyat"] = f"→ {clean['price']}"
        if dry_run:
            return "yeni", changes
        product = Product(
            tenant_id=store.tenant_id,
            brand_id=store.brand_id,
            sku=clean["sku"],
            name=clean["name"],
            vat_rate=clean["vat"],
        )
        session.add(product)
        session.flush()
        _write_versions(session, product, store, clean, today=today, user=user)
        return "yeni", changes

    if product.name != clean["name"]:
        changes["ad"] = f"{product.name} → {clean['name']}"
    if product.vat_rate != clean["vat"]:
        changes["kdv"] = f"%{product.vat_rate} → %{clean['vat']}"

    cost = _effective_cost(session, product.id)
    if clean["cost"] is not None and (cost is None or cost.unit_cost != clean["cost"]):
        changes["maliyet"] = f"{cost.unit_cost if cost else '—'} → {clean['cost']}"

    price = _effective_price(session, product.id, store.id)
    if clean["price"] is not None and (price is None or price.price != clean["price"]):
        changes["fiyat"] = f"{price.price if price else '—'} → {clean['price']}"

    logistics = _effective_logistics(session, product.id)
    if clean["desi"] is not None and (logistics is None or logistics.desi != clean["desi"]):
        changes["desi"] = f"{logistics.desi if logistics else '—'} → {clean['desi']}"

    if not changes:
        return "degisiklik_yok", {}
    if dry_run:
        return "guncelleme", changes

    product.name = clean["name"]
    product.vat_rate = clean["vat"]
    _write_versions(session, product, store, clean, today=today, user=user)
    return "guncelleme", changes


def _write_versions(
    session: Session,
    product: Product,
    store: Store,
    clean: dict[str, Any],
    *,
    today: date,
    user: str | None,
) -> None:
    """Maliyet/fiyat/desi değişikliklerini VERSİYONLU yazar — geçmiş kayıt ezilmez."""
    cost = _effective_cost(session, product.id)
    if clean["cost"] is not None and (cost is None or cost.unit_cost != clean["cost"]):
        session.add(
            SkuCost(
                product_id=product.id,
                unit_cost=clean["cost"],
                source=CostSource.MANUAL,
                effective_from=today,
                created_by=user,
            )
        )
    price = _effective_price(session, product.id, store.id)
    if clean["price"] is not None and (price is None or price.price != clean["price"]):
        session.add(
            SkuPrice(
                product_id=product.id,
                store_id=store.id,
                price=clean["price"],
                effective_from=today,
                created_by=user,
            )
        )
    logistics = _effective_logistics(session, product.id)
    if clean["desi"] is not None and (logistics is None or logistics.desi != clean["desi"]):
        session.add(SkuLogistics(product_id=product.id, desi=clean["desi"], effective_from=today))
    session.flush()


# Ürün seviyesinde olan, yani aynı SKU'nun tüm kanal satırlarında AYNI olması gereken
# alanlar. Fiyat kanal bazlıdır, bunlar değildir (spec §12A.1 sütun listesi).
PRODUCT_LEVEL_FIELDS = (("Alış Maliyeti", "cost"), ("Desi", "desi"), ("KDV %", "vat"))


def _conflicting_field(rows: list[tuple[ParsedRow, dict[str, Any]]]) -> str | None:
    """Aynı SKU'nun kanal satırları arasında ürün seviyesi alan çelişkisi var mı?

    Bir SKU birden fazla kanalda satılıyorsa dosyada birden fazla satırı olur; maliyet
    ve desi ürünün özelliğidir, satırdan satıra değişemez. Kullanıcı yalnızca bir satırı
    düzenlerse dosya kendi içinde çelişir — sessizce "son satır kazansın" demek veri
    kaybıdır, bu yüzden satırlar reddedilir (CLAUDE.md §5: en muhafazakâr seçenek).
    """
    for label, key in PRODUCT_LEVEL_FIELDS:
        values = {clean[key] for _, clean in rows if clean[key] is not None}
        if len(values) > 1:
            listed = ", ".join(str(value) for value in sorted(values))
            return f"Aynı SKU'nun satırlarında farklı {label}: {listed}"
    return None


def import_price_list(
    session: Session,
    payload: bytes,
    *,
    today: date,
    user: str | None = None,
    dry_run: bool = True,
    as_draft: bool = False,
    tenant_id: uuid.UUID | None = None,
    brand_id: uuid.UUID | None = None,
) -> ImportSummary:
    """Fiyat listesini işler. `dry_run=True` iken HİÇBİR yazma yapılmaz (spec §12A.2).

    `as_draft=True`: SKU'su boş satırlar ürün yerine **taslak** olarak alınır
    (spec §12A.3) — ürün ağacına yarım kayıt düşmez.
    """
    rows = parse_price_list(payload)
    stores = _stores_by_channel(session)
    summary = ImportSummary(dry_run=dry_run)

    def fail(row: ParsedRow, message: str) -> None:
        summary.hata += 1
        summary.rows.append(
            RowResult(
                row_no=row.row_no,
                sku=_text(row.values.get("SKU")),
                channel=_text(row.values.get("Kanal")),
                action="hata",
                message=message,
            )
        )

    # Önce doğrulama, sonra SKU bazında çelişki kontrolü, en son uygulama.
    validated: list[tuple[ParsedRow, dict[str, Any]]] = []
    for row in rows:
        clean, error = _validate(row, stores, as_draft=as_draft)
        if error:
            fail(row, error)
            continue
        # İzolasyon (spec §3A.2): başka markanın SKU'su bu markada ürün YARATMAZ.
        if (
            clean["sku"]
            and tenant_id is not None
            and brand_id is not None
            and belongs_to_another_brand(
                session, sku=clean["sku"], tenant_id=tenant_id, brand_id=brand_id
            )
        ):
            fail(row, CROSS_BRAND_MESSAGE)
            continue
        validated.append((row, clean))

    by_sku: dict[str, list[tuple[ParsedRow, dict[str, Any]]]] = {}
    for row, clean in validated:
        by_sku.setdefault(clean["sku"], []).append((row, clean))

    conflicted: set[int] = set()
    for sku, sku_rows in by_sku.items():
        if not sku:  # taslak adayları (SKU'suz) birbiriyle karşılaştırılmaz
            continue
        conflict = _conflicting_field(sku_rows) if len(sku_rows) > 1 else None
        if conflict:
            for row, _ in sku_rows:
                fail(row, conflict)
                conflicted.add(row.row_no)

    for row, clean in validated:
        if row.row_no in conflicted:
            continue
        if clean.get("is_draft"):
            summary.taslak += 1
            summary.rows.append(
                RowResult(
                    row_no=row.row_no,
                    sku="",
                    channel=clean["channel"],
                    action="taslak",
                    changes={"taslak": clean["name"]},
                )
            )
            if not dry_run and tenant_id is not None and brand_id is not None:
                _write_draft(session, clean, tenant_id=tenant_id, brand_id=brand_id)
            continue
        action, changes = _apply(
            session,
            clean,
            stores[clean["channel"]],
            today=today,
            user=user,
            dry_run=dry_run,
        )
        setattr(summary, action, getattr(summary, action) + 1)
        summary.rows.append(
            RowResult(
                row_no=row.row_no,
                sku=clean["sku"],
                channel=clean["channel"],
                action=action,
                changes=changes,
            )
        )

    log.info("pricelist.imported", rows=len(rows), **summary.as_dict())
    return summary


def _write_draft(
    session: Session,
    clean: dict[str, Any],
    *,
    tenant_id: uuid.UUID,
    brand_id: uuid.UUID,
) -> None:
    """SKU'suz satırı taslak ürün olarak kaydeder (spec §12A.3)."""
    session.add(
        ProductDraft(
            tenant_id=tenant_id,
            brand_id=brand_id,
            name=clean["name"],
            alis_maliyeti=clean["cost"] or ZERO,
            hedef_satis_fiyati=clean["price"] or ZERO,
            kanal=clean["channel"],
            vat_rate=clean["vat"],
            desi=clean["desi"],
            status=DraftStatus.DRAFT,
        )
    )
    session.flush()


def error_workbook(payload: bytes, summary: ImportSummary) -> bytes:
    """Orijinal dosya + "Hatalar" sayfası (satır no + açıklama) — spec §12A.2.3."""
    workbook = load_workbook(BytesIO(payload))
    if ERROR_SHEET_NAME in workbook.sheetnames:
        del workbook[ERROR_SHEET_NAME]
    sheet = workbook.create_sheet(ERROR_SHEET_NAME)
    sheet.append(("Satır", "SKU", "Kanal", "Hata"))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in summary.rows:
        if row.action == "hata":
            sheet.append((row.row_no, row.sku, row.channel, row.message))
    for index, width in enumerate((8, 18, 14, 70), start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
