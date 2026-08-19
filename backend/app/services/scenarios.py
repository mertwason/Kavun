"""Fiyat senaryosu motoru (spec §12A.4).

**Deterministik**: talep tahmini, elastikiyet, olasılık YOK. Girdi → hesap → sonuç.
Hesabın kendisi kâr motorundan gelir (`app/engine/profit.py`); bu katman senaryo
girdilerini motorun girdisine çevirir ve sonucu toplar.

Kargoyu satıcı ödemiyorsa (`alici`/`platform`) satıcının kargo maliyeti sıfırdır —
alıcının ödediği kargo satıcının geliri sayılmaz (muhafazakâr varsayım, CLAUDE.md §5).
"""

from __future__ import annotations

import uuid
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.logging import get_logger
from app.engine.pricing import PriceInputs, break_even_price, price_for_margin
from app.engine.profit import LineInput, ProfitBreakdown, compute_line_profit
from app.engine.vat import quantize_money
from app.models.catalog import Product, SkuCost
from app.models.enums import CommissionMode, CommissionSource, ShippingPayer
from app.models.identity import Store
from app.models.workspace import PricingScenario
from app.services.commission import resolve_commission

log = get_logger("services.scenarios")

ZERO = Decimal("0")
HUNDRED = Decimal("100")
MAX_COMPARE = 3
"""Spec §12A.4: karşılaştırmada en fazla 3 senaryo."""


class ScenarioError(RuntimeError):
    """Senaryo hesabının reddettiği durum."""


@dataclass(frozen=True)
class ScenarioResult:
    """Bir senaryonun sonucu (spec §12A.4)."""

    scenario_id: uuid.UUID | None
    name: str
    product_id: uuid.UUID
    sku: str
    satis_fiyati: Decimal
    """Liste fiyatı (kampanya indirimi öncesi)."""

    musteri_odedigi: Decimal
    """Kampanya sonrası müşterinin ödediği brüt tutar."""

    adet: int
    birim_kar: Decimal
    marj_pct: Decimal
    toplam_kar: Decimal
    basabas_fiyat: Decimal | None
    commission_rate: Decimal | None
    commission_source: CommissionSource | None
    cargo_cost: Decimal
    service_fee: Decimal
    breakdown: ProfitBreakdown


@dataclass(frozen=True)
class ScenarioInput:
    """Senaryo girdileri — kayıtlı senaryodan ya da formdan gelir."""

    name: str
    satis_fiyati: Decimal
    kampanya_indirim_pct: Decimal | None = None
    kampanya_satici_pay_pct: Decimal | None = None
    kargo_kim_oder: ShippingPayer = ShippingPayer.SATICI
    adet_varsayimi: int = 1
    commission_mode: CommissionMode = CommissionMode.CURRENT
    pinned_commission_rate: Decimal | None = None
    kargo_tahmini: Decimal | None = None
    scenario_id: uuid.UUID | None = None


def _effective_cost(session: Session, product_id: uuid.UUID, on_date: date) -> Decimal | None:
    cost = session.scalar(
        select(SkuCost)
        .where(SkuCost.product_id == product_id, SkuCost.effective_from <= on_date)
        .order_by(SkuCost.effective_from.desc(), SkuCost.created_at.desc())
        .limit(1)
    )
    return cost.unit_cost if cost else None


def _commission(
    session: Session,
    scenario: ScenarioInput,
    store: Store,
    product: Product,
    on_date: date,
) -> tuple[Decimal | None, CommissionSource | None]:
    """Senaryonun komisyon oranı: sabitlenmiş oran ya da tarifeden çözülen (spec §12B.4)."""
    if scenario.commission_mode is CommissionMode.PINNED and scenario.pinned_commission_rate:
        return scenario.pinned_commission_rate, CommissionSource.MANUAL
    rate_row, source = resolve_commission(
        session, store_id=store.id, product=product, on_date=on_date
    )
    return (rate_row.rate if rate_row else None), source


def evaluate(
    session: Session,
    *,
    product: Product,
    store: Store,
    scenario: ScenarioInput,
    on_date: date,
) -> ScenarioResult:
    """Senaryoyu hesaplar — kâr motorunu çağırır, ikinci bir formül yok."""
    unit_cost = _effective_cost(session, product.id, on_date)
    rate, source = _commission(session, scenario, store, product, on_date)

    # Kargoyu satıcı ödemiyorsa satıcının maliyeti yok (gelir de yazılmaz).
    cargo = (
        (scenario.kargo_tahmini or ZERO)
        if scenario.kargo_kim_oder is ShippingPayer.SATICI
        else ZERO
    )
    service_fee = store.service_fee_per_order or ZERO

    discount_rate = (scenario.kampanya_indirim_pct or ZERO) / HUNDRED
    seller_share = (
        (scenario.kampanya_satici_pay_pct / HUNDRED)
        if scenario.kampanya_satici_pay_pct is not None
        else Decimal("1")
    )
    discount_amount = quantize_money(scenario.satis_fiyati * discount_rate)
    customer_pays = quantize_money(scenario.satis_fiyati - discount_amount)

    breakdown = compute_line_profit(
        LineInput(
            line_gross=customer_pays,
            qty=1,
            vat_percent=product.vat_rate,
            unit_cost_net=unit_cost,
            commission_rate=rate,
            commission_source=source,
            cargo_cost=cargo,
            service_fee=service_fee,
            campaign_discount=discount_amount,
            campaign_seller_share_rate=seller_share,
        )
    )

    price_inputs = PriceInputs(
        unit_cost_net=unit_cost or ZERO,
        vat_percent=product.vat_rate,
        commission_rate=rate or ZERO,
        cargo_cost=cargo,
        service_fee=service_fee,
        campaign_discount_rate=discount_rate,
        campaign_seller_share_rate=seller_share,
    )
    adet = max(scenario.adet_varsayimi, 1)

    return ScenarioResult(
        scenario_id=scenario.scenario_id,
        name=scenario.name,
        product_id=product.id,
        sku=product.sku,
        satis_fiyati=quantize_money(scenario.satis_fiyati),
        musteri_odedigi=customer_pays,
        adet=adet,
        birim_kar=breakdown.profit,
        marj_pct=breakdown.margin_pct,
        toplam_kar=quantize_money(breakdown.profit * adet),
        basabas_fiyat=break_even_price(price_inputs) if unit_cost is not None else None,
        commission_rate=rate,
        commission_source=source,
        cargo_cost=cargo,
        service_fee=service_fee,
        breakdown=breakdown,
    )


def from_record(record: PricingScenario, *, cargo_estimate: Decimal | None = None) -> ScenarioInput:
    """Kayıtlı senaryoyu hesap girdisine çevirir.

    Kargo varsayımı senaryonun kendisinde saklanır; `cargo_estimate` yalnızca kayıt
    anında (henüz flush edilmemiş değeri geçmek için) kullanılır.
    """
    return ScenarioInput(
        name=record.name,
        satis_fiyati=record.satis_fiyati,
        kampanya_indirim_pct=record.kampanya_indirim_pct,
        kampanya_satici_pay_pct=record.kampanya_satici_pay_pct,
        kargo_kim_oder=record.kargo_kim_oder,
        adet_varsayimi=record.adet_varsayimi,
        commission_mode=record.commission_mode,
        pinned_commission_rate=record.pinned_commission_rate,
        kargo_tahmini=cargo_estimate if cargo_estimate is not None else record.kargo_tahmini,
        scenario_id=record.id,
    )


def compare(
    session: Session,
    *,
    scenarios: list[tuple[Product, Store, ScenarioInput]],
    on_date: date,
) -> list[ScenarioResult]:
    """En fazla 3 senaryoyu yan yana hesaplar (spec §12A.4)."""
    if len(scenarios) > MAX_COMPARE:
        raise ScenarioError(f"En fazla {MAX_COMPARE} senaryo karşılaştırılabilir")
    return [
        evaluate(session, product=product, store=store, scenario=scenario, on_date=on_date)
        for product, store, scenario in scenarios
    ]


@dataclass(frozen=True)
class TargetMarginResult:
    """Hedef marj çözücüsünün sonucu (spec §12A.4)."""

    target_margin_pct: Decimal
    price: Decimal | None
    reachable: bool
    result: ScenarioResult | None
    message: str = ""


def solve_target_margin(
    session: Session,
    *,
    product: Product,
    store: Store,
    target_margin_pct: Decimal,
    scenario: ScenarioInput,
    on_date: date,
) -> TargetMarginResult:
    """Hedef marjı tutturan minimum liste fiyatını KAPALI FORMÜLLE çözer (spec §12A.4).

    Sonuç, aynı motora geri verilerek doğrulanabilir olmalıdır (kabul kriteri §12A.6:
    ±0,01 puan). İterasyon yok.
    """
    unit_cost = _effective_cost(session, product.id, on_date)
    if unit_cost is None:
        return TargetMarginResult(
            target_margin_pct=target_margin_pct,
            price=None,
            reachable=False,
            result=None,
            message="Ürünün maliyeti tanımlı değil; hedef fiyat hesaplanamaz",
        )

    rate, _ = _commission(session, scenario, store, product, on_date)
    cargo = (
        (scenario.kargo_tahmini or ZERO)
        if scenario.kargo_kim_oder is ShippingPayer.SATICI
        else ZERO
    )
    discount_rate = (scenario.kampanya_indirim_pct or ZERO) / HUNDRED
    seller_share = (
        (scenario.kampanya_satici_pay_pct / HUNDRED)
        if scenario.kampanya_satici_pay_pct is not None
        else Decimal("1")
    )

    price = price_for_margin(
        target_margin_pct,
        PriceInputs(
            unit_cost_net=unit_cost,
            vat_percent=product.vat_rate,
            commission_rate=rate or ZERO,
            cargo_cost=cargo,
            service_fee=store.service_fee_per_order or ZERO,
            campaign_discount_rate=discount_rate,
            campaign_seller_share_rate=seller_share,
        ),
    )
    if price is None:
        return TargetMarginResult(
            target_margin_pct=target_margin_pct,
            price=None,
            reachable=False,
            result=None,
            message=("Bu maliyet ve komisyon yapısıyla hedef marja hiçbir fiyatta ulaşılamaz"),
        )

    solved = ScenarioInput(
        name=scenario.name,
        satis_fiyati=price,
        kampanya_indirim_pct=scenario.kampanya_indirim_pct,
        kampanya_satici_pay_pct=scenario.kampanya_satici_pay_pct,
        kargo_kim_oder=scenario.kargo_kim_oder,
        adet_varsayimi=scenario.adet_varsayimi,
        commission_mode=scenario.commission_mode,
        pinned_commission_rate=scenario.pinned_commission_rate,
        kargo_tahmini=scenario.kargo_tahmini,
    )
    return TargetMarginResult(
        target_margin_pct=target_margin_pct,
        price=price,
        reachable=True,
        result=evaluate(session, product=product, store=store, scenario=solved, on_date=on_date),
    )


# --- senaryo xlsx round-trip (spec §12A.4) ----------------------------------

SCENARIO_TEMPLATE_VERSION = "kavun-senaryo-v1"
SCENARIO_SHEET = "Senaryolar"
SCENARIO_COLUMNS = (
    "SKU",
    "Senaryo Adı",
    "Satış Fiyatı",
    "Kampanya İndirimi %",
    "Kampanya Satıcı Payı %",
    "Kargo Kim Öder",
    "Kargo Tahmini",
    "Adet Varsayımı",
    # Aşağıdakiler hesap sonucudur; import'ta yok sayılır (kaynak motordur).
    "Birim Kâr",
    "Marj %",
    "Toplam Kâr",
    "Başabaş Fiyat",
)
SCENARIO_COMPUTED_COLUMNS = ("Birim Kâr", "Marj %", "Toplam Kâr", "Başabaş Fiyat")


def _scenario_rows(payload: bytes) -> list[dict[str, Any]]:
    """Senaryo dosyasını okur ve şablon sürümünü doğrular."""
    try:
        workbook = load_workbook(BytesIO(payload), data_only=True)
    except Exception as exc:  # openpyxl çeşitli hata tipleri fırlatır
        raise ScenarioError("Dosya okunamadı; Excel (.xlsx) bekleniyor.") from exc

    sheet = workbook[SCENARIO_SHEET] if SCENARIO_SHEET in workbook.sheetnames else workbook.active
    if sheet is None or str(sheet["A1"].value or "").strip() != SCENARIO_TEMPLATE_VERSION:
        raise ScenarioError(f"Şablon sürümü uyumsuz. Beklenen: {SCENARIO_TEMPLATE_VERSION}.")
    headers = [str(cell.value or "").strip() for cell in sheet[2]]
    missing = [name for name in SCENARIO_COLUMNS if name not in headers]
    if missing:
        raise ScenarioError(f"Eksik sütun(lar): {', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    for values in sheet.iter_rows(min_row=3, values_only=True):
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        rows.append(
            {
                name: values[headers.index(name)]
                for name in SCENARIO_COLUMNS
                if headers.index(name) < len(values)
            }
        )
    return rows


def _decimal_or_none(value: Any) -> Decimal | None:
    if value is None or str(value).strip() == "":
        return None
    return Decimal(str(value).replace(",", ".").strip())  # allow-float: openpyxl hücresi


def build_scenario_workbook(results: list[ScenarioResult], *, inputs: list[ScenarioInput]) -> bytes:
    """Senaryoları hesaplanmış sonuç sütunlarıyla birlikte xlsx olarak yazar (spec §12A.4)."""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = SCENARIO_SHEET
    sheet["A1"] = SCENARIO_TEMPLATE_VERSION
    sheet.row_dimensions[1].hidden = True

    for index, name in enumerate(SCENARIO_COLUMNS, start=1):
        cell = sheet.cell(row=2, column=index, value=name)
        cell.font = Font(bold=True)
        sheet.column_dimensions[get_column_letter(index)].width = 20
    sheet.freeze_panes = "A3"

    for row_no, (result, scenario) in enumerate(zip(results, inputs, strict=True), start=3):
        values = (
            result.sku,
            result.name,
            result.satis_fiyati,
            scenario.kampanya_indirim_pct,
            scenario.kampanya_satici_pay_pct,
            scenario.kargo_kim_oder.value,
            scenario.kargo_tahmini,
            result.adet,
            result.birim_kar,
            result.marj_pct,
            result.toplam_kar,
            result.basabas_fiyat,
        )
        for index, value in enumerate(values, start=1):
            sheet.cell(row=row_no, column=index, value=value)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_scenario_workbook(
    session: Session, payload: bytes
) -> list[tuple[Product, ScenarioInput]]:
    """Dosyadaki senaryoları okur; hesap sonucu sütunları YOK SAYILIR (spec §12A.4)."""
    prepared: list[tuple[Product, ScenarioInput]] = []
    for index, row in enumerate(_scenario_rows(payload), start=1):
        sku = str(row.get("SKU") or "").strip()
        product = session.scalar(select(Product).where(Product.sku == sku)) if sku else None
        if product is None:
            raise ScenarioError(f"{index}. senaryo: bilinmeyen SKU '{sku or '(boş)'}'")

        price = _decimal_or_none(row.get("Satış Fiyatı"))
        if price is None or price <= ZERO:
            raise ScenarioError(f"{index}. senaryo: Satış Fiyatı zorunlu ve pozitif olmalı")

        payer_raw = str(row.get("Kargo Kim Öder") or ShippingPayer.SATICI.value).strip().lower()
        try:
            payer = ShippingPayer(payer_raw)
        except ValueError as exc:
            allowed = ", ".join(item.value for item in ShippingPayer)
            raise ScenarioError(
                f"{index}. senaryo: geçersiz 'Kargo Kim Öder': {payer_raw} (geçerli: {allowed})"
            ) from exc

        quantity = _decimal_or_none(row.get("Adet Varsayımı"))
        prepared.append(
            (
                product,
                ScenarioInput(
                    name=str(row.get("Senaryo Adı") or f"Senaryo {index}").strip(),
                    satis_fiyati=price,
                    kampanya_indirim_pct=_decimal_or_none(row.get("Kampanya İndirimi %")),
                    kampanya_satici_pay_pct=_decimal_or_none(row.get("Kampanya Satıcı Payı %")),
                    kargo_kim_oder=payer,
                    adet_varsayimi=int(quantity) if quantity else 1,
                    kargo_tahmini=_decimal_or_none(row.get("Kargo Tahmini")),
                ),
            )
        )
    return prepared
