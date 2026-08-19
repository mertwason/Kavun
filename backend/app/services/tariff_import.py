"""Komisyon tarifesi Excel yüklemesi — esnek parser (spec §12B.2).

> Kullanıcı Trendyol'dan indirdiği dosyayı **formatını değiştirmeden** yükleyebilmelidir.

Bu yüzden sabit sütun sırası varsayılmaz. Parser:
1. başlık satırını arar (ilk 20 satır içinde, tanınan başlık sayısı en yüksek olan),
2. sütunları Türkçe başlık varyasyonlarıyla **fuzzy** eşleştirir
   (`Komisyon %`, `Komisyon Oranı`, `Kategori`, `Ana Kategori`, `Alt Kategori`, …),
3. eşleştirmeyi dry-run yanıtında kullanıcıya gösterir: "şu sütunu kategori, şu sütunu
   oran olarak okudum — onaylıyor musun?"

Çok seviyeli kategori (ana > alt > yaprak) desteklenir; eşleştirme **en spesifik**
seviyeden yapılır. Kavun'daki ürün kategorileriyle eşleşmeyen tarife satırları hata
DEĞİLDİR — `unmatched` listesinde raporlanır (ileride o kategoride ürün açılırsa kayıt
zaten yerinde olur).

Oran yazımı serbesttir: `%21,5`, `21,5`, `0,215` ve `21.5` aynı orana çözülür — Excel'de
hücre bazen metin bazen sayı gelir, dosya "olduğu gibi" yüklenebilmelidir.
"""

from __future__ import annotations

import re
import unicodedata
import uuid
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.logging import get_logger
from app.models.catalog import CommissionRate, Product
from app.models.enums import CommissionScope, CommissionSource
from app.models.identity import Store
from app.services import tariffs

log = get_logger("services.tariff_import")

ZERO = Decimal("0")
ONE = Decimal("1")
HUNDRED = Decimal("100")

HEADER_SEARCH_ROWS = 20
"""Başlık satırı dosyanın ilk satırı olmak zorunda değil (logo/başlık blokları olur)."""

# Başlık sözlüğü. Kanal bazlı sözlükler Faz 3'te buraya eklenecek (Hepsiburada/N11).
CATEGORY_HEADERS = (
    "kategori",
    "kategori adi",
    "kategori ismi",
    "urun kategorisi",
    "ana kategori",
    "ust kategori",
    "alt kategori",
    "yaprak kategori",
    "kategori kirilimi",
    "category",
)
RATE_HEADERS = (
    "komisyon",
    "komisyon %",
    "komisyon orani",
    "komisyon oran",
    "komisyon yuzdesi",
    "oran",
    "commission",
    "commission rate",
)
CODE_HEADERS = ("kategori kodu", "kod", "category code", "categoryid", "kategori id")
CAMPAIGN_HEADERS = ("kampanya", "kampanya donemi", "kampanyali komisyon", "kampanya orani")

# Kategori seviyeleri en genelden en spesifiğe; eşleştirme sondan başlar.
LEVEL_ORDER = ("ana kategori", "ust kategori", "kategori", "alt kategori", "yaprak kategori")


class TariffFileError(ValueError):
    """Dosya okunamadı ya da beklenen sütunlar bulunamadı."""


def normalize(text: Any) -> str:
    """Başlık karşılaştırması için sadeleştirme: küçük harf, aksansız, tek boşluk."""
    raw = str(text or "").strip().lower()
    raw = raw.replace("ı", "i").replace("İ", "i")
    decomposed = unicodedata.normalize("NFKD", raw)
    stripped = "".join(char for char in decomposed if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", stripped.replace("%", " % ")).strip()


def _matches(header: str, candidates: tuple[str, ...]) -> bool:
    """Başlık, adaylardan birini içeriyor mu (kısmi eşleşme yeterli)."""
    normalized = normalize(header)
    if not normalized:
        return False
    return any(normalize(candidate) in normalized for candidate in candidates)


@dataclass
class ColumnMapping:
    """Parser'ın hangi sütunu ne olarak okuduğu — dry-run'da kullanıcıya gösterilir."""

    header_row: int
    rate_column: int
    rate_header: str
    category_columns: list[tuple[int, str]] = field(default_factory=list)
    code_column: int | None = None
    code_header: str | None = None
    campaign_column: int | None = None
    campaign_header: str | None = None

    def as_dict(self) -> dict[str, Any]:
        """UI'ın "şunu şöyle okudum" kutusu için."""
        return {
            "header_row": self.header_row,
            "rate_header": self.rate_header,
            "category_headers": [header for _, header in self.category_columns],
            "code_header": self.code_header,
            "campaign_header": self.campaign_header,
        }


@dataclass
class TariffRow:
    """Dosyadan okunan bir tarife satırı."""

    row_no: int
    category: str
    """En spesifik kategori seviyesi (eşleştirme bunun üzerinden yapılır)."""

    category_path: list[str]
    category_code: str | None
    rate: Decimal
    is_campaign_period: bool


@dataclass
class ImportPreview:
    """Dry-run yanıtı: eşleştirme + fark analizi (spec §12B.2)."""

    mapping: dict[str, Any]
    valid_from: date
    total_rows: int = 0
    matched: int = 0
    unchanged: int = 0
    changed: int = 0
    new_categories: int = 0
    unmatched: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    changes: list[dict[str, Any]] = field(default_factory=list)
    affected_sku_count: int = 0
    monthly_profit_impact: Decimal = ZERO
    written: int = 0

    def as_dict(self) -> dict[str, Any]:
        """Log dostu özet."""
        return {
            "total_rows": self.total_rows,
            "matched": self.matched,
            "changed": self.changed,
            "new": self.new_categories,
            "unmatched": len(self.unmatched),
            "errors": len(self.errors),
            "written": self.written,
        }


def parse_rate(value: Any) -> Decimal | None:
    """`%21,5` · `21,5` · `0,215` · `21.5` → `Decimal("0.2150")`.

    Sezgi: 1'den büyük değerler yüzde, 1'den küçük/eşit değerler zaten orandır.
    `%1` gibi tek haneli yüzdeler de doğru okunur çünkü `%` işareti korunur.
    """
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    has_percent = "%" in text
    text = text.replace("%", "").strip().replace(" ", "")
    # "21,5" (TR ondalık) ve "1.234,5" (TR binlik) biçimlerini normalize et.
    if "," in text and "." in text:
        text = text.replace(".", "")
    text = text.replace(",", ".")
    try:
        number = Decimal(text)
    except InvalidOperation:
        return None
    if number < ZERO:
        return None
    if has_percent or number > ONE:
        number = number / HUNDRED
    return number.quantize(Decimal("0.0001"))


def detect_mapping(sheet: Any) -> ColumnMapping:
    """Başlık satırını bulur ve sütunları eşleştirir (spec §12B.2 esnek parser)."""
    best: ColumnMapping | None = None
    best_score = 0

    for row_no, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=HEADER_SEARCH_ROWS, values_only=True), start=1
    ):
        headers = [str(value or "").strip() for value in row]
        rate_column: int | None = None
        rate_header = ""
        categories: list[tuple[int, str]] = []
        code_column: int | None = None
        code_header: str | None = None
        campaign_column: int | None = None
        campaign_header: str | None = None

        for index, header in enumerate(headers):
            if not header:
                continue
            if rate_column is None and _matches(header, RATE_HEADERS):
                if _matches(header, CAMPAIGN_HEADERS):
                    campaign_column, campaign_header = index, header
                    continue
                rate_column, rate_header = index, header
                continue
            if _matches(header, CODE_HEADERS):
                code_column, code_header = index, header
                continue
            if _matches(header, CATEGORY_HEADERS):
                categories.append((index, header))
                continue
            if _matches(header, CAMPAIGN_HEADERS):
                campaign_column, campaign_header = index, header

        score = (2 if rate_column is not None else 0) + len(categories)
        if rate_column is not None and categories and score > best_score:
            best_score = score
            best = ColumnMapping(
                header_row=row_no,
                rate_column=rate_column,
                rate_header=rate_header,
                category_columns=categories,
                code_column=code_column,
                code_header=code_header,
                campaign_column=campaign_column,
                campaign_header=campaign_header,
            )

    if best is None:
        raise TariffFileError(
            "Kategori ve komisyon oranı sütunları bulunamadı. "
            "Dosyada 'Kategori' ve 'Komisyon %' benzeri başlıklar olmalı."
        )
    # Seviye sırası: en spesifik sütun sona gelsin (eşleştirme sondan yapılır).
    best.category_columns.sort(key=lambda item: _level_rank(item[1]))
    return best


def _level_rank(header: str) -> int:
    """Kategori sütununun seviye sırası; tanınmayan başlık ortada sayılır."""
    normalized = normalize(header)
    for rank, level in enumerate(LEVEL_ORDER):
        if normalize(level) in normalized:
            return rank
    return len(LEVEL_ORDER) // 2


def parse_workbook(payload: bytes) -> tuple[ColumnMapping, list[TariffRow], list[str]]:
    """Dosyayı okur; `(eşleştirme, satırlar, hatalar)` döner."""
    try:
        workbook = load_workbook(BytesIO(payload), data_only=True)
    except Exception as exc:  # openpyxl çeşitli hata tipleri fırlatır
        raise TariffFileError("Dosya okunamadı; Excel (.xlsx) bekleniyor.") from exc

    sheet = workbook.active
    if sheet is None:
        raise TariffFileError("Dosyada sayfa bulunamadı.")

    mapping = detect_mapping(sheet)
    rows: list[TariffRow] = []
    errors: list[str] = []

    for row_no, values in enumerate(
        sheet.iter_rows(min_row=mapping.header_row + 1, values_only=True),
        start=mapping.header_row + 1,
    ):
        if all(value is None or str(value).strip() == "" for value in values):
            continue

        path = [
            str(values[index]).strip()
            for index, _ in mapping.category_columns
            if index < len(values) and values[index] is not None and str(values[index]).strip()
        ]
        if not path:
            continue

        raw_rate = values[mapping.rate_column] if mapping.rate_column < len(values) else None
        if raw_rate is None or str(raw_rate).strip() == "":
            # Oran hücresi tamamen boş: bu bir veri satırı değil (dosya sonundaki notlar,
            # ara başlıklar). Hata saymak kullanıcıyı sahte hatalarla boğardı.
            continue
        rate = parse_rate(raw_rate)
        if rate is None:
            errors.append(f"{row_no}. satır: komisyon oranı okunamadı ({raw_rate!r})")
            continue
        if rate > ONE:
            errors.append(f"{row_no}. satır: komisyon oranı %100'den büyük ({raw_rate!r})")
            continue

        code = None
        if mapping.code_column is not None and mapping.code_column < len(values):
            code_value = values[mapping.code_column]
            code = str(code_value).strip() if code_value is not None else None

        campaign = False
        if mapping.campaign_column is not None and mapping.campaign_column < len(values):
            campaign_value = values[mapping.campaign_column]
            campaign = bool(campaign_value) and str(campaign_value).strip().lower() not in (
                "0",
                "hayır",
                "hayir",
                "false",
                "yok",
                "-",
            )

        rows.append(
            TariffRow(
                row_no=row_no,
                category=path[-1],
                category_path=path,
                category_code=code,
                rate=rate,
                is_campaign_period=campaign,
            )
        )

    return mapping, rows, errors


def _known_categories(session: Session) -> dict[str, str]:
    """Kavun'daki ürün kategorileri: normalize → orijinal."""
    values = session.scalars(select(Product.category).where(Product.category.is_not(None))).all()
    return {normalize(value): value for value in values if value}


def _match_category(row: TariffRow, known: dict[str, str]) -> str | None:
    """En spesifik seviyeden başlayarak Kavun kategorisiyle eşleştirir."""
    for candidate in reversed(row.category_path):
        matched = known.get(normalize(candidate))
        if matched:
            return matched
    # Tam yol da denenir: "Kahve > Harman" biçimindeki dosyalar için.
    joined = normalize("/".join(row.category_path))
    return known.get(joined)


def import_tariff(
    session: Session,
    payload: bytes,
    *,
    store: Store,
    valid_from: date,
    today: date,
    dry_run: bool = True,
    user: str | None = None,
) -> ImportPreview:
    """Tarife dosyasını işler; dry-run'da yazma YOK, fark analizi döner (spec §12B.2)."""
    mapping, rows, errors = parse_workbook(payload)
    known = _known_categories(session)

    preview = ImportPreview(
        mapping=mapping.as_dict(),
        valid_from=valid_from,
        total_rows=len(rows),
        errors=errors,
    )

    changed: list[tuple[Product, Decimal, Decimal]] = []
    for row in rows:
        category = _match_category(row, known)
        if category is None:
            preview.unmatched.append(row.category)
            continue

        preview.matched += 1
        current = tariffs._effective_rate(
            session, store_id=store.id, category=category, on_date=today
        )
        if current is None:
            preview.new_categories += 1
        elif current.rate == row.rate:
            preview.unchanged += 1
            continue
        else:
            preview.changed += 1
            preview.changes.append(
                {
                    "category": category,
                    "old_rate": str(current.rate),
                    "new_rate": str(row.rate),
                }
            )
            changed.extend(
                (product, current.rate, row.rate)
                for product in session.scalars(
                    select(Product).where(Product.category == category)
                ).all()
            )

        if not dry_run:
            session.add(
                CommissionRate(
                    store_id=store.id,
                    scope=CommissionScope.CATEGORY,
                    category_code=category,
                    rate=row.rate,
                    source=CommissionSource.MANUAL_TARIFF_UPLOAD,
                    valid_from=valid_from,
                    is_campaign_period=row.is_campaign_period,
                )
            )
            preview.written += 1

    # Otomatik fark analizi: "bu tarife sana ne yapacak" (spec §12B.2).
    impact = tariffs.estimate_impact(session, store=store, on_date=today, changes=changed)
    preview.affected_sku_count = impact.affected_sku_count
    preview.monthly_profit_impact = impact.monthly_profit_impact

    if not dry_run:
        session.flush()

    log.info("tariff_import.processed", store_id=str(store.id), user=user, **preview.as_dict())
    return preview


def future_rate(
    session: Session,
    *,
    store_id: uuid.UUID,
    category: str | None,
    on_date: date,
) -> CommissionRate | None:
    """İleri tarihli tarife — senaryoların `future_tariff` modu bunu kullanır (§12B.4)."""
    if not category:
        return None
    candidates = list(
        session.scalars(
            select(CommissionRate).where(
                CommissionRate.store_id == store_id,
                CommissionRate.scope == CommissionScope.CATEGORY,
                CommissionRate.category_code == category,
                CommissionRate.valid_from <= on_date,
                (CommissionRate.valid_to.is_(None)) | (CommissionRate.valid_to > on_date),
            )
        ).all()
    )
    if not candidates:
        return None
    return max(candidates, key=lambda item: (item.valid_from, item.created_at))
