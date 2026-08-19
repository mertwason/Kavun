"""Kargo faturası eşleştirme: `estimated → actual` (spec §5.3, §6.2).

Faz 1'de kargo maliyeti **tahminidir** (desi × tarife). Gerçek tutar ay sonunda kargo
firmasının faturasından gelir; bu modül o faturayı okuyup gönderilerle eşleştirir ve
maliyeti kesinleştirir. Kesinleşen her satır kâr motorunu yeniden tetikler ve değişen
alanlar `profit_revisions`'a append-only loglanır (spec §6.2).

## Neden xlsx

Kargo firmaları (Aras, Yurtiçi, Trendyol Express) fatura dökümünü Excel olarak verir;
API'leri satıcı tarafına açık değildir. Şablon disiplini fiyat listesindekiyle aynıdır:
**indirilen dosya = yüklenen şablon**, sürüm hücresi tutmazsa dosya reddedilir.

## Eşleştirme anahtarı

Öncelik **gönderi (takip) numarasıdır**; kanal onu vermiyorsa **sipariş numarası**
kullanılır. İkisi de tutmazsa satır "eşleşmedi" kuyruğuna düşer — uydurma eşleştirme
yapılmaz, çünkü yanlış gönderiye yazılan maliyet sessizce yanlış kâr üretir.

## Kesinleşmiş maliyet ezilmez

`cost_state = actual` olan gönderi ikinci bir faturayla güncellenmez; düzeltme gerekiyorsa
yeni fatura satırı "zaten kesinleşmiş" olarak raporlanır. Bu, normalize akışındaki kuralın
(kesinleşmiş maliyet tahminle ezilmez) aynısıdır.
"""

from __future__ import annotations

import uuid
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.logging import get_logger
from app.engine.vat import quantize_money
from app.models.enums import AlertSeverity, CostState
from app.models.identity import Store
from app.models.results import Alert
from app.models.transactions import CargoInvoice, Order, Shipment

log = get_logger("services.cargo")

ZERO = Decimal("0")

TEMPLATE_VERSION = "kavun-kargo-v1"
META_CELL = "A1"
HEADER_ROW = 2
FIRST_DATA_ROW = 3
SHEET_NAME = "Kargo Faturası"

COLUMNS = ("Sipariş No", "Gönderi No", "Tarih", "Desi", "Tutar")
COLUMN_WIDTHS = (24, 24, 14, 10, 14)

UNMATCHED_ALERT = "kargo_faturasi_eslesmedi"


class TemplateError(RuntimeError):
    """Şablon okunamadı ya da sürümü uyuşmuyor."""


@dataclass
class RowResult:
    """Bir fatura satırının sonucu."""

    row_no: int
    reference: str
    action: str
    """`kesinlesti` · `zaten_kesin` · `eslesmedi` · `hata`"""

    amount: Decimal = ZERO
    previous: Decimal | None = None
    message: str = ""

    def as_dict(self) -> dict[str, Any]:
        """JSON gösterimi (fatura kaydının `lines` alanına da bu yazılır)."""
        return {
            "row_no": self.row_no,
            "reference": self.reference,
            "action": self.action,
            "amount": str(self.amount),
            "previous": None if self.previous is None else str(self.previous),
            "message": self.message,
        }


@dataclass
class MatchSummary:
    """Yükleme özeti — `dry_run` ile gerçek koşuda aynı yapı."""

    dry_run: bool = True
    rows: int = 0
    kesinlesti: int = 0
    zaten_kesin: int = 0
    eslesmedi: int = 0
    hata: int = 0
    total_amount: Decimal = ZERO
    delta: Decimal = ZERO
    """Tahmin ile gerçek arasındaki fark toplamı (pozitif = tahmin düşük kalmış)."""

    invoice_id: uuid.UUID | None = None
    results: list[RowResult] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        """JSON gösterimi."""
        return {
            "dry_run": self.dry_run,
            "rows": self.rows,
            "kesinlesti": self.kesinlesti,
            "zaten_kesin": self.zaten_kesin,
            "eslesmedi": self.eslesmedi,
            "hata": self.hata,
            "total_amount": str(self.total_amount),
            "delta": str(self.delta),
            "results": [row.as_dict() for row in self.results],
        }


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _decimal(value: Any, field_name: str) -> Decimal:
    """Hücreyi `Decimal`e çevirir; Türkçe ondalık ayracını kabul eder."""
    if value is None or _text(value) == "":
        raise ValueError(f"{field_name} boş olamaz")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):  # allow-float: openpyxl hücreyi float döndürebilir
        return Decimal(str(value))
    try:
        return Decimal(_text(value).replace(".", "").replace(",", "."))
    except InvalidOperation as exc:
        raise ValueError(f"{field_name} sayı değil") from exc


def _as_date(value: Any) -> date:
    """Hücreden tarih çıkarır."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValueError("Tarih okunamadı (GG.AA.YYYY bekleniyor)")


def template_workbook() -> bytes:
    """Boş kargo faturası şablonu."""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = SHEET_NAME
    sheet[META_CELL] = TEMPLATE_VERSION
    sheet.row_dimensions[1].hidden = True

    for index, name in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=HEADER_ROW, column=index, value=name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[get_column_letter(index)].width = COLUMN_WIDTHS[index - 1]
    sheet.freeze_panes = f"A{FIRST_DATA_ROW}"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_workbook(payload: bytes) -> list[tuple[int, dict[str, Any]]]:
    """Dosyayı okur, şablon sürümünü doğrular; hücreleri henüz yorumlamaz."""
    try:
        workbook = load_workbook(BytesIO(payload), data_only=True)
    except Exception as exc:  # openpyxl çeşitli hata tipleri fırlatır
        raise TemplateError("Dosya okunamadı; Excel (.xlsx) bekleniyor.") from exc

    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
    if sheet is None:
        raise TemplateError("Dosyada sayfa bulunamadı.")
    if _text(sheet[META_CELL].value) != TEMPLATE_VERSION:
        raise TemplateError(
            f"Şablon sürümü uyumsuz. Beklenen: {TEMPLATE_VERSION}. "
            "Güncel şablonu 'Şablonu indir' ile alın."
        )

    headers = [_text(cell.value) for cell in sheet[HEADER_ROW]]
    missing = [name for name in COLUMNS if name not in headers]
    if missing:
        raise TemplateError(f"Eksik sütun(lar): {', '.join(missing)}")

    rows: list[tuple[int, dict[str, Any]]] = []
    for row_no, row in enumerate(
        sheet.iter_rows(min_row=FIRST_DATA_ROW, values_only=True), start=FIRST_DATA_ROW
    ):
        if all(value is None or _text(value) == "" for value in row):
            continue
        values = {
            name: row[headers.index(name)] for name in COLUMNS if headers.index(name) < len(row)
        }
        rows.append((row_no, values))
    return rows


def _find_shipment(
    session: Session, store: Store, *, tracking: str, order_no: str
) -> Shipment | None:
    """Gönderiyi önce takip numarasından, sonra sipariş numarasından bulur."""
    if tracking:
        shipment = session.scalar(select(Shipment).where(Shipment.tracking_no == tracking))
        if shipment is not None:
            return shipment
    if order_no:
        order = session.scalar(
            select(Order).where(Order.store_id == store.id, Order.external_order_id == order_no)
        )
        if order is not None:
            return session.scalar(select(Shipment).where(Shipment.order_id == order.id))
    return None


def import_invoice(
    session: Session,
    *,
    payload: bytes,
    store: Store,
    invoice_no: str,
    period: str,
    dry_run: bool = True,
) -> MatchSummary:
    """Kargo faturasını gönderilerle eşleştirir ve maliyeti kesinleştirir.

    `dry_run=True` iken hiçbir şey yazılmaz; sayımlar ve eşleşmeyen satırlar aynı biçimde
    döner. Uygulandığında kesinleşen siparişlerin kârı yeniden hesaplanır ve değişen
    alanlar `profit_revisions`'a loglanır (spec §6.2).
    """
    from app.services.profit import recompute_orders

    summary = MatchSummary(dry_run=dry_run)
    parsed = parse_workbook(payload)
    summary.rows = len(parsed)

    touched_orders: list[uuid.UUID] = []

    for row_no, values in parsed:
        tracking = _text(values.get("Gönderi No"))
        order_no = _text(values.get("Sipariş No"))
        reference = tracking or order_no or "—"
        try:
            _as_date(values.get("Tarih"))
            amount = _decimal(values.get("Tutar"), "Tutar")
            desi = _decimal(values.get("Desi"), "Desi") if _text(values.get("Desi")) else None
        except ValueError as exc:
            summary.hata += 1
            summary.results.append(
                RowResult(row_no=row_no, reference=reference, action="hata", message=str(exc))
            )
            continue

        if amount < ZERO:
            summary.hata += 1
            summary.results.append(
                RowResult(
                    row_no=row_no,
                    reference=reference,
                    action="hata",
                    message="Tutar negatif olamaz",
                )
            )
            continue

        shipment = _find_shipment(session, store, tracking=tracking, order_no=order_no)
        if shipment is None:
            summary.eslesmedi += 1
            summary.results.append(
                RowResult(
                    row_no=row_no,
                    reference=reference,
                    action="eslesmedi",
                    amount=amount,
                    message="Gönderi bulunamadı",
                )
            )
            continue

        if shipment.cost_state is CostState.ACTUAL:
            summary.zaten_kesin += 1
            summary.results.append(
                RowResult(
                    row_no=row_no,
                    reference=reference,
                    action="zaten_kesin",
                    amount=amount,
                    previous=shipment.cargo_cost_actual,
                    message="Bu gönderinin maliyeti zaten kesinleşmiş",
                )
            )
            continue

        estimated = shipment.cargo_cost_estimated
        summary.kesinlesti += 1
        summary.total_amount += amount
        summary.delta += amount - estimated
        summary.results.append(
            RowResult(
                row_no=row_no,
                reference=reference,
                action="kesinlesti",
                amount=amount,
                previous=estimated,
            )
        )

        if dry_run:
            continue

        shipment.cargo_cost_actual = quantize_money(amount)
        shipment.desi_invoiced = desi
        shipment.cost_state = CostState.ACTUAL
        if tracking and not shipment.tracking_no:
            shipment.tracking_no = tracking
        touched_orders.append(shipment.order_id)

    if dry_run:
        return summary

    invoice = CargoInvoice(
        tenant_id=store.tenant_id,
        brand_id=store.brand_id,
        store_id=store.id,
        invoice_no=invoice_no,
        period=period,
        total=quantize_money(summary.total_amount),
        lines=[row.as_dict() for row in summary.results],
    )
    session.add(invoice)
    session.flush()
    summary.invoice_id = invoice.id

    if summary.eslesmedi:
        session.add(
            Alert(
                tenant_id=store.tenant_id,
                brand_id=store.brand_id,
                type=UNMATCHED_ALERT,
                severity=AlertSeverity.WARNING,
                entity_ref=f"cargo_invoice:{invoice_no}",
                message=(
                    f"{invoice_no} faturasında {summary.eslesmedi} satır gönderiyle "
                    "eşleşmedi; maliyet tahmini kaldı."
                ),
                created_at=datetime.now(UTC),
            )
        )

    session.flush()
    if touched_orders:
        # Kesinleşen maliyet kârı değiştirir; revizyonlar `profit_revisions`'a düşer (§6.2).
        recompute_orders(session, order_ids=touched_orders, reason="kargo_faturasi")

    log.info(
        "cargo.invoice_imported",
        invoice_no=invoice_no,
        kesinlesti=summary.kesinlesti,
        eslesmedi=summary.eslesmedi,
        delta=str(summary.delta),
    )
    return summary


def invoices(session: Session) -> list[CargoInvoice]:
    """Marka kapsamlı kargo faturaları (en yeni üstte)."""
    return list(
        session.scalars(select(CargoInvoice).order_by(CargoInvoice.created_at.desc())).all()
    )


@dataclass(frozen=True)
class CostStateSummary:
    """Kargo maliyetinin kesinleşme durumu — dashboard rozetinin veri kaynağı."""

    total: int
    actual: int
    estimated: int
    estimated_amount: Decimal
    actual_amount: Decimal


def cost_state_summary(session: Session) -> CostStateSummary:
    """Kaç gönderinin maliyeti kesinleşti, kaçı hâlâ tahmini."""
    shipments = list(session.scalars(select(Shipment)).all())
    actual = [item for item in shipments if item.cost_state is CostState.ACTUAL]
    estimated = [item for item in shipments if item.cost_state is not CostState.ACTUAL]
    return CostStateSummary(
        total=len(shipments),
        actual=len(actual),
        estimated=len(estimated),
        estimated_amount=quantize_money(
            sum((item.cargo_cost_estimated for item in estimated), ZERO)
        ),
        actual_amount=quantize_money(
            sum((item.cargo_cost_actual or ZERO for item in actual), ZERO)
        ),
    )
