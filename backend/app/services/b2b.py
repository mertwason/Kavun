"""D2B / kurumsal satış kanalı (spec §12C.9).

Alessi'nin kurumsal satışları pazaryerinden gelmez: elle ya da xlsx ile girilir. Kavun
bunları **normal sipariş** olarak yazar — böylece stok düşer, kâr motoru aynı formülle
hesaplar ve satış marka P&L'ine dahil olur. Tek fark kanaldır:

- **Komisyon 0** ve pazaryeri hizmet bedeli yok (satış pazaryerinde olmadı).
- Fiyat müşteri kademesine göre iskontoludur; iskonto satır fiyatına yansır ve
  "hangi kademe ne marj bırakıyor" analizi `customers.tier` üzerinden yapılır.

Şablon disiplini KVN-10'daki ile aynıdır: **indirilen dosya = yüklenen şablon**, sürüm
hücresi uyuşmazsa dosya reddedilir. Yükleme `dry_run` ile önizlenir; onaya kadar hiçbir
sipariş yazılmaz.

İdempotency: sipariş numarası satırdan türetilir (`D2B-<tarih>-<müşteri>-<no>`); aynı
dosya iki kez yüklenirse sipariş çoğalmaz, satır "zaten var" olarak raporlanır.
"""

from __future__ import annotations

import hashlib
import uuid
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.logging import get_logger
from app.engine.vat import quantize_money
from app.models.catalog import Customer, Product
from app.models.enums import ChannelCode, CommissionSource, OrderStatus
from app.models.identity import Channel, Store
from app.models.transactions import Order, OrderLine

log = get_logger("services.b2b")

ZERO = Decimal("0")
HUNDRED = Decimal("100")

TEMPLATE_VERSION = "kavun-d2b-v1"
META_CELL = "A1"
HEADER_ROW = 2
FIRST_DATA_ROW = 3
SHEET_NAME = "D2B Satışlar"

COLUMNS = ("Tarih", "Müşteri", "Kademe", "SKU", "Adet", "Birim Fiyat", "İskonto %", "KDV %")
COLUMN_WIDTHS = (14, 32, 12, 18, 8, 14, 12, 8)

VALID_VAT_RATES = (Decimal("0"), Decimal("1"), Decimal("10"), Decimal("20"))


class TemplateError(RuntimeError):
    """Şablon okunamadı ya da sürümü uyuşmuyor."""


@dataclass(frozen=True)
class ParsedRow:
    """Ham satır — henüz doğrulanmadı."""

    row_no: int
    values: dict[str, Any]


@dataclass
class RowError:
    """Reddedilen satır ve gerekçesi."""

    row_no: int
    sku: str
    reason: str

    def as_dict(self) -> dict[str, Any]:
        """JSON gösterimi."""
        return {"row_no": self.row_no, "sku": self.sku, "reason": self.reason}


@dataclass
class ImportSummary:
    """Yükleme özeti (dry-run ve gerçek koşuda aynı yapı)."""

    rows: int = 0
    orders: int = 0
    lines: int = 0
    customers: int = 0
    skipped: int = 0
    """Zaten yazılmış (aynı dosya ikinci kez yüklendi) satır sayısı."""

    gross_total: Decimal = ZERO
    errors: list[RowError] = field(default_factory=list)

    def as_dict(self) -> dict[str, Any]:
        """JSON gösterimi."""
        return {
            "rows": self.rows,
            "orders": self.orders,
            "lines": self.lines,
            "customers": self.customers,
            "skipped": self.skipped,
            "gross_total": str(self.gross_total),
            "errors": [error.as_dict() for error in self.errors],
        }


@dataclass(frozen=True)
class TierMargin:
    """Müşteri kademesi bazlı satış özeti (spec §12C.9 "hangi tier ne marj bırakıyor")."""

    tier: str
    customers: int
    orders: int
    qty: int
    revenue: Decimal
    avg_discount_pct: Decimal


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _decimal(value: Any, field_name: str) -> Decimal:
    """Hücreyi `Decimal`e çevirir; virgüllü yazımı kabul eder."""
    if value is None or _text(value) == "":
        raise ValueError(f"{field_name} boş olamaz")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, float):  # allow-float: openpyxl hücreyi float döndürebilir
        return Decimal(str(value))
    try:
        return Decimal(_text(value).replace(".", "").replace(",", "."))
    except InvalidOperation as exc:
        raise ValueError(f"{field_name} sayı değil") from exc


def _as_date(value: Any) -> date:
    """Hücreden tarih çıkarır (`datetime`, `date` ya da `GG.AA.YYYY` metni)."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValueError("Tarih okunamadı (GG.AA.YYYY bekleniyor)")


def template_workbook() -> bytes:
    """Boş D2B şablonu — indirilen dosya birebir yüklenebilir olmalıdır."""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = SHEET_NAME
    sheet[META_CELL] = TEMPLATE_VERSION
    sheet.row_dimensions[1].hidden = True

    for index, name in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=HEADER_ROW, column=index, value=name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[get_column_letter(index)].width = COLUMN_WIDTHS[index - 1]
    sheet.freeze_panes = f"A{FIRST_DATA_ROW}"

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def parse_workbook(payload: bytes) -> list[ParsedRow]:
    """Dosyayı okur ve şablon sürümünü doğrular; hücreleri henüz yorumlamaz."""
    try:
        workbook = load_workbook(BytesIO(payload), data_only=True)
    except Exception as exc:  # openpyxl çeşitli hata tipleri fırlatır
        raise TemplateError("Dosya okunamadı; Excel (.xlsx) bekleniyor.") from exc

    sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
    if sheet is None:
        raise TemplateError("Dosyada sayfa bulunamadı.")
    if _text(sheet[META_CELL].value) != TEMPLATE_VERSION:
        raise TemplateError(
            f"Şablon sürümü uyumsuz. Beklenen: {TEMPLATE_VERSION}. "
            "Güncel şablonu 'Şablonu indir' ile alın."
        )

    headers = [_text(cell.value) for cell in sheet[HEADER_ROW]]
    missing = [name for name in COLUMNS if name not in headers]
    if missing:
        raise TemplateError(f"Eksik sütun(lar): {', '.join(missing)}")

    rows: list[ParsedRow] = []
    for row_no, row in enumerate(
        sheet.iter_rows(min_row=FIRST_DATA_ROW, values_only=True), start=FIRST_DATA_ROW
    ):
        if all(value is None or _text(value) == "" for value in row):
            continue
        values = {
            name: row[headers.index(name)] for name in COLUMNS if headers.index(name) < len(row)
        }
        rows.append(ParsedRow(row_no=row_no, values=values))
    return rows


def _external_order_id(sale_date: date, customer: str) -> str:
    """Sipariş numarası satırdan türetilir: aynı gün aynı müşteri = aynı sipariş.

    Böylece bir dosya iki kez yüklendiğinde sipariş çoğalmaz (spec §3.7 idempotency).
    """
    digest = hashlib.sha1(customer.strip().lower().encode("utf-8")).hexdigest()[:8]
    return f"D2B-{sale_date:%Y%m%d}-{digest}"


def d2b_store(session: Session) -> Store | None:
    """Aktif markanın manuel (D2B) mağazası."""
    channel = session.scalar(select(Channel).where(Channel.code == ChannelCode.MANUAL))
    if channel is None:
        return None
    return session.scalar(select(Store).where(Store.channel_id == channel.id))


def import_sales(
    session: Session, *, payload: bytes, store: Store, dry_run: bool = True
) -> ImportSummary:
    """xlsx satışlarını sipariş olarak yazar (spec §12C.9).

    `dry_run=True` iken hiçbir şey yazılmaz; sayımlar ve hatalar aynı biçimde döner —
    kullanıcı önizlemeyi görüp onaylar.
    """
    summary = ImportSummary()
    parsed = parse_workbook(payload)
    summary.rows = len(parsed)

    products = {product.sku: product for product in session.scalars(select(Product)).all()}
    customers = {
        customer.name.strip().lower(): customer
        for customer in session.scalars(select(Customer)).all()
    }
    orders: dict[str, Order] = {}

    for row in parsed:
        sku = _text(row.values.get("SKU"))
        try:
            sale_date = _as_date(row.values.get("Tarih"))
            customer_name = _text(row.values.get("Müşteri"))
            if not customer_name:
                raise ValueError("Müşteri boş olamaz")
            qty = int(_decimal(row.values.get("Adet"), "Adet"))
            unit_price = _decimal(row.values.get("Birim Fiyat"), "Birim Fiyat")
            discount = (
                _decimal(row.values.get("İskonto %"), "İskonto %")
                if _text(row.values.get("İskonto %"))
                else ZERO
            )
            vat_rate = _decimal(row.values.get("KDV %"), "KDV %")
        except ValueError as exc:
            summary.errors.append(RowError(row_no=row.row_no, sku=sku, reason=str(exc)))
            continue

        product = products.get(sku)
        if product is None:
            summary.errors.append(
                RowError(row_no=row.row_no, sku=sku, reason="SKU bu markada bulunamadı")
            )
            continue
        if qty <= 0:
            summary.errors.append(
                RowError(row_no=row.row_no, sku=sku, reason="Adet pozitif olmalı")
            )
            continue
        if unit_price <= ZERO:
            summary.errors.append(
                RowError(row_no=row.row_no, sku=sku, reason="Birim fiyat pozitif olmalı")
            )
            continue
        if not ZERO <= discount < HUNDRED:
            summary.errors.append(
                RowError(row_no=row.row_no, sku=sku, reason="İskonto 0 ile 100 arasında olmalı")
            )
            continue
        if vat_rate not in VALID_VAT_RATES:
            summary.errors.append(
                RowError(row_no=row.row_no, sku=sku, reason=f"Geçersiz KDV oranı: {vat_rate}")
            )
            continue

        net_price = quantize_money(unit_price * (HUNDRED - discount) / HUNDRED)
        line_gross = quantize_money(net_price * qty)
        summary.gross_total += line_gross
        summary.lines += 1

        if dry_run:
            continue

        key = customer_name.strip().lower()
        customer = customers.get(key)
        if customer is None:
            customer = Customer(
                tenant_id=store.tenant_id,
                brand_id=store.brand_id,
                name=customer_name,
                tier=_text(row.values.get("Kademe")) or None,
                default_discount_pct=discount,
            )
            session.add(customer)
            session.flush()
            customers[key] = customer
            summary.customers += 1

        external_id = _external_order_id(sale_date, customer_name)
        order = orders.get(external_id)
        if order is None:
            order = session.scalar(
                select(Order).where(
                    Order.store_id == store.id, Order.external_order_id == external_id
                )
            )
            if order is None:
                order = Order(
                    tenant_id=store.tenant_id,
                    brand_id=store.brand_id,
                    store_id=store.id,
                    customer_id=customer.id,
                    external_order_id=external_id,
                    order_date=datetime.combine(sale_date, datetime.min.time(), tzinfo=UTC),
                    status=OrderStatus.DELIVERED,
                    gross_total=ZERO,
                    currency="TRY",
                )
                session.add(order)
                session.flush()
                summary.orders += 1
            orders[external_id] = order

        line_id = f"{sku}-{row.row_no}"
        exists = session.scalar(
            select(OrderLine).where(
                OrderLine.order_id == order.id, OrderLine.external_line_id == line_id
            )
        )
        if exists is not None:
            summary.skipped += 1
            summary.lines -= 1
            summary.gross_total -= line_gross
            continue

        session.add(
            OrderLine(
                tenant_id=store.tenant_id,
                brand_id=store.brand_id,
                order_id=order.id,
                product_id=product.id,
                external_line_id=line_id,
                qty=qty,
                unit_sale_price=net_price,
                line_gross=line_gross,
                vat_rate=vat_rate,
                # D2B satışı pazaryerinde olmadı: komisyon YOK (spec §12C.9).
                commission_rate_used=ZERO,
                commission_source=CommissionSource.MANUAL,
                status=OrderStatus.DELIVERED,
            )
        )
        order.gross_total = quantize_money(order.gross_total + line_gross)

    if not dry_run:
        session.flush()
    log.info(
        "b2b.sales_imported", dry_run=dry_run, **{"rows": summary.rows, "lines": summary.lines}
    )
    return summary


def tier_margins(session: Session) -> list[TierMargin]:
    """Kademe bazlı satış özeti — hangi kademe ne bırakıyor (spec §12C.9).

    Sipariş → müşteri bağı `orders.customer_id` üzerindedir; kademesi girilmemiş
    müşteriler "—" kademesinde toplanır.
    """
    store = d2b_store(session)
    if store is None:
        return []

    orders = list(session.scalars(select(Order).where(Order.store_id == store.id)).all())
    if not orders:
        return []

    customers = {customer.id: customer for customer in session.scalars(select(Customer)).all()}
    order_ids = {order.id for order in orders}
    tier_of_order = {
        order.id: (
            customers[order.customer_id].tier or "—" if order.customer_id in customers else "—"
        )
        for order in orders
    }

    revenue: dict[str, Decimal] = {}
    qty: dict[str, int] = {}
    order_count: dict[str, set[uuid.UUID]] = {}
    for line in session.scalars(select(OrderLine)).all():
        if line.order_id not in order_ids or line.status is OrderStatus.CANCELLED:
            continue
        tier = tier_of_order[line.order_id]
        revenue[tier] = revenue.get(tier, ZERO) + line.line_gross
        qty[tier] = qty.get(tier, 0) + line.qty
        order_count.setdefault(tier, set()).add(line.order_id)

    by_tier: dict[str, list[Customer]] = {}
    for customer in customers.values():
        by_tier.setdefault(customer.tier or "—", []).append(customer)

    result: list[TierMargin] = []
    for tier in sorted(set(revenue) | set(by_tier)):
        members = by_tier.get(tier, [])
        discounts = [item.default_discount_pct or ZERO for item in members]
        result.append(
            TierMargin(
                tier=tier,
                customers=len(members),
                orders=len(order_count.get(tier, set())),
                qty=qty.get(tier, 0),
                revenue=quantize_money(revenue.get(tier, ZERO)),
                avg_discount_pct=(
                    (sum(discounts, ZERO) / Decimal(len(discounts))).quantize(Decimal("0.01"))
                    if discounts
                    else ZERO
                ),
            )
        )
    return result
