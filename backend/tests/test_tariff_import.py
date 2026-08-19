"""KVN-14: tarife Excel yüklemesi — esnek parser (spec §12B.2).

Kabul kriterleri (§12B.2):
- Tarife dosyası SIFIR manuel müdahale ile okunur (başlık satırı ilk satırda değil,
  sütun sırası sabit değil, oranlar metin)
- İleri tarihli yükleme + `future_tariff` senaryosu uçtan uca çalışır
- Eşleşmeyen kategoriler HATA değil, `unmatched` listesi
"""

from __future__ import annotations

from collections.abc import Iterator
from datetime import timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api import deps
from app.core.context import RequestContext, system_scope, use_context
from app.main import create_app
from app.models.catalog import CommissionRate, Product, SkuPrice
from app.models.enums import CommissionMode, CommissionSource, UserRole
from app.models.identity import Brand, Store
from app.services import profit as profit_service
from app.services import scenarios, tariff_import
from tests.profit_factories import ORDER_DATE, make_commission, make_order, make_product, make_store

D = Decimal
TODAY = ORDER_DATE.date()
FIXTURE = Path(__file__).parent / "fixtures" / "tariffs" / "trendyol_komisyon_2026_09.xlsx"


@pytest.fixture
def payload() -> bytes:
    """Kanalın yayımladığı biçimi taklit eden tarife dosyası."""
    return FIXTURE.read_bytes()


@pytest.fixture
def store(db_session: Session) -> Iterator[Store]:
    """Mağaza + marka bağlamı."""
    with system_scope():
        store = make_store(db_session)
        store.service_fee_per_order = D("12.0000")
        db_session.flush()
        brand = db_session.get(Brand, store.brand_id)
    assert brand is not None
    context = RequestContext(
        tenant_id=brand.tenant_id,
        user_id=None,
        brand_id=brand.id,
        brand_slug=brand.slug,
        role=UserRole.ADMIN,
    )
    with use_context(context):
        yield store


@pytest.fixture
def product(db_session: Session, store: Store) -> Product:
    """Kahve/Harman kategorisinde, satışı olan ürün."""
    product = make_product(
        db_session, store, "TARIFE-IMP-1", cost=D("50.0000"), category="Kahve/Harman"
    )
    db_session.add(
        SkuPrice(
            product_id=product.id,
            store_id=store.id,
            price=D("199.0000"),
            effective_from=TODAY - timedelta(days=60),
        )
    )
    db_session.flush()
    return product


# --- oran ayrıştırma (spec §12B.2) ------------------------------------------


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("%21,5", D("0.2150")),
        ("21,5", D("0.2150")),
        ("21.5", D("0.2150")),
        ("0,215", D("0.2150")),
        (0.215, D("0.2150")),
        ("%1", D("0.0100")),
        ("%0", D("0.0000")),
        (" %14,50 ", D("0.1450")),
    ],
)
def test_rate_is_read_in_every_common_notation(raw: Any, expected: Decimal) -> None:
    """Dosya "olduğu gibi" yüklenebilmeli: oran metin de sayı da olabilir."""
    assert tariff_import.parse_rate(raw) == expected


@pytest.mark.parametrize("raw", ["", None, "abc", "-5"])
def test_unreadable_rate_returns_none(raw: Any) -> None:
    """Okunamayan oran uydurulmaz."""
    assert tariff_import.parse_rate(raw) is None


# --- esnek parser (spec §12B.2 kabul kriteri) -------------------------------


def test_fixture_is_read_without_manual_intervention(payload: bytes) -> None:
    """§12B.2 kabul: başlık satırı ilk satırda değil, sütun sırası farklı — yine de okunur."""
    mapping, rows, errors = tariff_import.parse_workbook(payload)

    assert mapping.header_row == 4  # üstteki duyuru blokları atlandı
    assert "Komisyon" in mapping.rate_header
    assert errors == []
    assert len(rows) == 14
    harman = next(row for row in rows if row.category == "Kahve/Harman")
    assert harman.rate == D("0.1600")
    assert harman.category_path == ["Gıda", "Kahve/Harman"]
    assert harman.category_code == "1001"


def test_mapping_is_reported_for_user_confirmation(payload: bytes) -> None:
    """§12B.2: "şu sütunu kategori, şu sütunu oran olarak okudum" bilgisi döner."""
    mapping, _, _ = tariff_import.parse_workbook(payload)

    reported = mapping.as_dict()
    assert reported["rate_header"] == "Komisyon Oranı (%)"
    assert reported["category_headers"] == ["Ana Kategori", "Alt Kategori"]
    assert reported["code_header"] == "Kategori Kodu"


def test_campaign_column_is_recognised(payload: bytes) -> None:
    """Kampanya dönemi sütunu tanınır ve işaretli satır kampanya kaydı olur."""
    _, rows, _ = tariff_import.parse_workbook(payload)

    dekorasyon = next(row for row in rows if row.category == "Dekorasyon")
    assert dekorasyon.is_campaign_period is True
    assert all(not row.is_campaign_period for row in rows if row.category != "Dekorasyon")


def test_header_variations_are_matched() -> None:
    """Türkçe başlık varyasyonları desteklenir (spec §12B.2)."""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(("Kategori", "Komisyon %"))
    sheet.append(("Kahve/Harman", "16"))
    buffer = BytesIO()
    workbook.save(buffer)

    mapping, rows, errors = tariff_import.parse_workbook(buffer.getvalue())

    assert mapping.header_row == 1
    assert errors == []
    assert rows[0].rate == D("0.1600")


def test_file_without_recognisable_headers_is_rejected() -> None:
    """Kategori/oran sütunu bulunamazsa anlaşılır hata verilir."""
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(("Foo", "Bar"))
    sheet.append(("x", "y"))
    buffer = BytesIO()
    workbook.save(buffer)

    with pytest.raises(tariff_import.TariffFileError, match="bulunamadı"):
        tariff_import.parse_workbook(buffer.getvalue())


def test_non_excel_payload_is_rejected() -> None:
    """Excel olmayan dosya anlaşılır hata verir."""
    with pytest.raises(tariff_import.TariffFileError, match="okunamadı"):
        tariff_import.parse_workbook(b"bu bir excel degil")


# --- eşleştirme ve import (spec §12B.2) -------------------------------------


def test_unmatched_categories_are_reported_not_failed(
    db_session: Session, store: Store, product: Product, payload: bytes
) -> None:
    """§12B.2: Kavun'da karşılığı olmayan tarife satırı HATA değildir."""
    preview = tariff_import.import_tariff(
        db_session,
        payload,
        store=store,
        valid_from=TODAY,
        today=TODAY,
        dry_run=True,
    )

    assert preview.errors == []
    assert "Telefon/Aksesuar" in preview.unmatched
    assert preview.matched >= 1


def test_dry_run_writes_nothing_but_reports_impact(
    db_session: Session, store: Store, product: Product, payload: bytes
) -> None:
    """§12B.2: dry-run'da yazma yok ama "bu tarife sana ne yapacak" raporu var."""
    make_commission(
        db_session,
        store,
        rate=D("0.1450"),
        category="Kahve/Harman",
        valid_from=ORDER_DATE - timedelta(days=60),
    )
    order = make_order(db_session, store, [(product, 10, D("1990.0000"))])
    profit_service.recompute_orders(db_session, order_ids=[order.id])
    before = len(db_session.scalars(select(CommissionRate)).all())

    preview = tariff_import.import_tariff(
        db_session, payload, store=store, valid_from=TODAY, today=TODAY, dry_run=True
    )

    assert preview.written == 0
    assert len(db_session.scalars(select(CommissionRate)).all()) == before
    assert preview.changed >= 1  # %14,5 → %16,0
    assert preview.monthly_profit_impact < 0
    assert preview.affected_sku_count >= 1
    change = next(item for item in preview.changes if item["category"] == "Kahve/Harman")
    assert change["old_rate"] == "0.1450"
    assert change["new_rate"] == "0.1600"


def test_apply_writes_rates_with_upload_source(
    db_session: Session, store: Store, product: Product, payload: bytes
) -> None:
    """§12B.2: kayıtlar `manual_tariff_upload` kaynağıyla yazılır."""
    preview = tariff_import.import_tariff(
        db_session, payload, store=store, valid_from=TODAY, today=TODAY, dry_run=False
    )

    assert preview.written >= 1
    written = db_session.scalars(
        select(CommissionRate).where(CommissionRate.source == CommissionSource.MANUAL_TARIFF_UPLOAD)
    ).all()
    assert len(written) == preview.written
    assert all(rate.valid_from == TODAY for rate in written)


def test_unchanged_rates_are_not_rewritten(
    db_session: Session, store: Store, product: Product, payload: bytes
) -> None:
    """Aynı oran yeniden yazılmaz — tarife tablosu gereksiz kayıtla şişmez."""
    make_commission(
        db_session,
        store,
        rate=D("0.1600"),
        category="Kahve/Harman",
        valid_from=ORDER_DATE - timedelta(days=60),
    )

    preview = tariff_import.import_tariff(
        db_session, payload, store=store, valid_from=TODAY, today=TODAY, dry_run=True
    )

    assert preview.unchanged >= 1
    assert not any(item["category"] == "Kahve/Harman" for item in preview.changes)


# --- ileri tarihli yükleme + future_tariff senaryosu (spec §12B.2, §12B.4) --


def test_future_dated_upload_does_not_change_today(
    db_session: Session, store: Store, product: Product, payload: bytes
) -> None:
    """§12B.2: ileri tarihli tarife bugünün hesabını ETKİLEMEZ."""
    make_commission(
        db_session,
        store,
        rate=D("0.1450"),
        category="Kahve/Harman",
        valid_from=ORDER_DATE - timedelta(days=60),
    )
    future = TODAY + timedelta(days=30)

    tariff_import.import_tariff(
        db_session, payload, store=store, valid_from=future, today=TODAY, dry_run=False
    )

    from app.services.commission import resolve_commission

    resolved, _ = resolve_commission(db_session, store_id=store.id, product=product, on_date=TODAY)
    assert resolved is not None
    assert resolved.rate == D("0.1450")  # bugün hâlâ eski oran


def test_future_tariff_scenario_uses_the_announced_rate(
    db_session: Session, store: Store, product: Product, payload: bytes
) -> None:
    """§12B.4 + §12B.2 kabul: `future_tariff` modu duyurulan oranı kullanır."""
    make_commission(
        db_session,
        store,
        rate=D("0.1450"),
        category="Kahve/Harman",
        valid_from=ORDER_DATE - timedelta(days=60),
    )
    future = TODAY + timedelta(days=30)
    tariff_import.import_tariff(
        db_session, payload, store=store, valid_from=future, today=TODAY, dry_run=False
    )

    current = scenarios.evaluate(
        db_session,
        product=product,
        store=store,
        scenario=scenarios.ScenarioInput(name="Bugün", satis_fiyati=D("199.00")),
        on_date=TODAY,
    )
    announced = scenarios.evaluate(
        db_session,
        product=product,
        store=store,
        scenario=scenarios.ScenarioInput(
            name="Yeni tarife",
            satis_fiyati=D("199.00"),
            commission_mode=CommissionMode.FUTURE_TARIFF,
            future_tariff_date=future,
        ),
        on_date=TODAY,
    )

    assert current.commission_rate == D("0.1450")
    assert announced.commission_rate == D("0.1600")
    assert announced.birim_kar < current.birim_kar


# --- API katmanı -------------------------------------------------------------


@pytest.fixture
def api(db_session: Session, store: Store, product: Product) -> Iterator[TestClient]:
    """Test oturumuna bağlı API istemcisi."""

    async def session_override() -> Any:
        yield db_session

    app = create_app()
    app.dependency_overrides[deps.get_session] = session_override
    with TestClient(app, raise_server_exceptions=False) as client:
        yield client


def _headers(api: TestClient, brand: str = "alessi") -> dict[str, str]:
    response = api.post("/auth/dev-login", json={"email": "mert@mokkalabs.com", "brand": brand})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def test_upload_endpoint_dry_run_returns_mapping_and_impact(
    api: TestClient, payload: bytes
) -> None:
    """§12B.2: dry-run yanıtı eşleştirmeyi ve fark analizini taşır."""
    response = api.post(
        "/alessi/tariffs/upload",
        params={"valid_from": str(TODAY), "dry_run": True},
        files={"file": ("tarife.xlsx", payload)},
        headers=_headers(api),
    )

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["dry_run"] is True
    assert body["written"] == 0
    assert body["mapping"]["rate_header"] == "Komisyon Oranı (%)"
    assert body["unmatched"]


def test_upload_endpoint_applies_and_logs_batch(
    api: TestClient, db_session: Session, payload: bytes
) -> None:
    """Onaydan sonra kayıtlar yazılır ve import batch'e loglanır."""
    from app.models.workspace import ImportBatch

    response = api.post(
        "/alessi/tariffs/upload",
        params={"valid_from": str(TODAY), "dry_run": False},
        files={"file": ("tarife.xlsx", payload)},
        headers=_headers(api),
    )

    assert response.status_code == 200, response.text
    assert response.json()["written"] >= 1
    batch = db_session.scalar(select(ImportBatch).where(ImportBatch.kind == "commission_tariff"))
    assert batch is not None
    assert batch.filename == "tarife.xlsx"


def test_upload_endpoint_rejects_unreadable_file(api: TestClient) -> None:
    """Okunamayan dosya 422 + anlaşılır mesaj."""
    response = api.post(
        "/alessi/tariffs/upload",
        params={"valid_from": str(TODAY)},
        files={"file": ("bozuk.xlsx", b"excel degil")},
        headers=_headers(api),
    )

    assert response.status_code == 422
    assert "okunamadı" in response.json()["detail"]


def test_upload_requires_valid_from(api: TestClient, payload: bytes) -> None:
    """§12B.2: `valid_from` zorunlu — tarifenin ne zaman yürürlüğe girdiği bilinmeden yazılmaz."""
    response = api.post(
        "/alessi/tariffs/upload",
        files={"file": ("tarife.xlsx", payload)},
        headers=_headers(api),
    )

    assert response.status_code == 422


def test_fixture_stays_readable(payload: bytes) -> None:
    """Fixture bozulursa kabul kriteri sessizce kaybolmasın."""
    sheet = load_workbook(BytesIO(payload)).active
    assert sheet is not None
    assert sheet.cell(row=4, column=1).value == "Ana Kategori"
