"""KVN-18: D2B kanal, fire/hasar ve MSRP disiplini (spec §12C.9-10).

Kabul kriterleri (§12C.11):
- D2B satış importu → stok düşer, komisyon 0, marka P&L'de doğru kanalda görünür
- Damage hareketi → stok ve fire gideri doğru, **ortalama maliyet değişmez**
"""

from __future__ import annotations

from collections.abc import Iterator
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api import b2b as api_b2b
from app.api import deps
from app.core.context import RequestContext, system_scope, use_context
from app.main import create_app
from app.models.catalog import Customer, Product, SkuPrice
from app.models.enums import ChannelCode, InventoryMovement, OrderStatus, UserRole
from app.models.identity import Channel, Store
from app.models.inventory import InventoryLedger, SkuCostState
from app.models.results import Alert
from app.models.transactions import Order, OrderLine
from app.services import b2b, discipline, inventory
from tests.profit_factories import make_product, make_store

D = Decimal
SALE_DATE = date(2026, 8, 10)


@pytest.fixture
def store(db_session: Session) -> Iterator[Store]:
    """Alessi mağazası + marka bağlamı (D2B ve MSRP bayrakları açık markadır)."""
    from app.models.identity import Brand

    with system_scope():
        store = make_store(db_session)
        brand = db_session.get(Brand, store.brand_id)
    assert brand is not None
    context = RequestContext(
        tenant_id=brand.tenant_id,
        user_id=None,
        brand_id=brand.id,
        brand_slug=brand.slug,
        role=UserRole.ADMIN,
    )
    with use_context(context):
        yield store


@pytest.fixture
def d2b(db_session: Session, store: Store) -> Store:
    """Manuel kanal mağazası — `seed_base` kurar."""
    record = b2b.d2b_store(db_session)
    assert record is not None, "seed_base D2B mağazasını kurmalı"
    return record


@pytest.fixture
def product(db_session: Session, store: Store) -> Product:
    """Stoklu ürün."""
    record = make_product(db_session, store, "ALS-9090-3", cost=D("3000.0000"))
    inventory.opening_stock(
        db_session, product=record, qty=D("50"), unit_cost=D("3000"), on_date=SALE_DATE
    )
    return record


def _workbook(rows: list[tuple[Any, ...]]) -> bytes:
    """Şablonu doldurup baytlara çevirir."""
    workbook = load_workbook(BytesIO(b2b.template_workbook()))
    sheet = workbook[b2b.SHEET_NAME]
    for index, row in enumerate(rows, start=b2b.FIRST_DATA_ROW):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=index, column=column, value=value)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# --- şablon disiplini --------------------------------------------------------


def test_template_can_be_uploaded_back(db_session: Session, d2b: Store) -> None:
    """İndirilen dosya = yüklenen şablon (KVN-10 disiplini)."""
    summary = b2b.import_sales(db_session, payload=b2b.template_workbook(), store=d2b)

    assert summary.rows == 0
    assert summary.errors == []


def test_wrong_template_version_is_rejected(db_session: Session, d2b: Store) -> None:
    """Başka bir Excel yüklenirse akış durur — sessizce yanlış sütun okunmaz."""
    workbook = Workbook()
    buffer = BytesIO()
    workbook.save(buffer)

    with pytest.raises(b2b.TemplateError, match="Şablon sürümü"):
        b2b.import_sales(db_session, payload=buffer.getvalue(), store=d2b)


# --- D2B satış importu (kabul §12C.11) ---------------------------------------


def test_dry_run_writes_nothing(db_session: Session, d2b: Store, product: Product) -> None:
    """Önizleme sayar ama yazmaz."""
    payload = _workbook([(SALE_DATE, "Kurumsal A.Ş.", "Bayi", product.sku, 4, 8000, 10, 20)])

    summary = b2b.import_sales(db_session, payload=payload, store=d2b, dry_run=True)

    assert summary.lines == 1
    assert summary.orders == 0
    assert db_session.scalar(select(Order).where(Order.store_id == d2b.id)) is None


def test_import_creates_orders_without_commission(
    db_session: Session, d2b: Store, product: Product
) -> None:
    """§12C.11: D2B satışı sipariş olur, komisyon 0'dır."""
    payload = _workbook([(SALE_DATE, "Kurumsal A.Ş.", "Bayi", product.sku, 4, 8000, 10, 20)])

    summary = b2b.import_sales(db_session, payload=payload, store=d2b, dry_run=False)

    assert summary.orders == 1
    assert summary.lines == 1
    assert summary.customers == 1

    line = db_session.scalar(select(OrderLine))
    assert line is not None
    # 8.000 × %10 iskonto = 7.200
    assert line.unit_sale_price == D("7200.0000")
    assert line.line_gross == D("28800.0000")
    assert line.commission_rate_used == D("0")


def test_imported_sale_reduces_stock(db_session: Session, d2b: Store, product: Product) -> None:
    """§12C.11: D2B satışı stoktan düşer, ortalama maliyet korunur."""
    payload = _workbook([(SALE_DATE, "Kurumsal A.Ş.", "Bayi", product.sku, 4, 8000, 10, 20)])
    b2b.import_sales(db_session, payload=payload, store=d2b, dry_run=False)

    inventory.record_sales(db_session)

    state = db_session.scalar(select(SkuCostState).where(SkuCostState.product_id == product.id))
    assert state is not None
    assert state.on_hand_qty == D("46")
    assert state.avg_cost == D("3000.000000")


def test_import_is_idempotent(db_session: Session, d2b: Store, product: Product) -> None:
    """Aynı dosya iki kez yüklenirse sipariş çoğalmaz (spec §3.7)."""
    payload = _workbook([(SALE_DATE, "Kurumsal A.Ş.", "Bayi", product.sku, 4, 8000, 10, 20)])
    b2b.import_sales(db_session, payload=payload, store=d2b, dry_run=False)

    second = b2b.import_sales(db_session, payload=payload, store=d2b, dry_run=False)

    assert second.skipped == 1
    assert second.lines == 0
    assert len(db_session.scalars(select(Order).where(Order.store_id == d2b.id)).all()) == 1


def test_unknown_sku_is_reported_not_guessed(
    db_session: Session, d2b: Store, product: Product
) -> None:
    """Bilinmeyen SKU satırı reddedilir; diğer satırlar işlenir."""
    payload = _workbook(
        [
            (SALE_DATE, "Kurumsal A.Ş.", "Bayi", "YOK-123", 1, 100, 0, 20),
            (SALE_DATE, "Kurumsal A.Ş.", "Bayi", product.sku, 2, 8000, 0, 20),
        ]
    )

    summary = b2b.import_sales(db_session, payload=payload, store=d2b, dry_run=False)

    assert summary.lines == 1
    assert [error.reason for error in summary.errors] == ["SKU bu markada bulunamadı"]


def test_invalid_vat_rate_is_rejected(db_session: Session, d2b: Store, product: Product) -> None:
    """Türkiye'de geçerli olmayan KDV oranı satırı reddeder."""
    payload = _workbook([(SALE_DATE, "Kurumsal A.Ş.", "Bayi", product.sku, 1, 8000, 0, 18)])

    summary = b2b.import_sales(db_session, payload=payload, store=d2b, dry_run=False)

    assert summary.lines == 0
    assert "Geçersiz KDV" in summary.errors[0].reason


def test_tier_margins_group_by_customer_tier(
    db_session: Session, d2b: Store, product: Product
) -> None:
    """§12C.9: hangi kademe ne bırakıyor."""
    payload = _workbook(
        [
            (SALE_DATE, "Kurumsal A.Ş.", "Bayi", product.sku, 4, 8000, 10, 20),
            (SALE_DATE, "Otel B", "Anahtar", product.sku, 2, 8000, 20, 20),
        ]
    )
    b2b.import_sales(db_session, payload=payload, store=d2b, dry_run=False)

    rows = {row.tier: row for row in b2b.tier_margins(db_session)}

    assert rows["Bayi"].revenue == D("28800.0000")
    assert rows["Anahtar"].revenue == D("12800.0000")
    assert rows["Anahtar"].avg_discount_pct == D("20.00")


# --- fire/hasar (spec §12C.10, kabul §12C.11) --------------------------------


def test_damage_reduces_stock_without_changing_average(
    db_session: Session, product: Product
) -> None:
    """§12C.11: hasar hareketi stoku düşürür, **ortalama maliyet değişmez**."""
    entry = inventory.damage(db_session, product=product, qty=D("3"), reason="Kırıldı — depo")

    assert entry.movement is InventoryMovement.DAMAGE
    assert entry.qty_delta == D("-3")
    assert entry.on_hand_after == D("47")
    assert entry.avg_cost_after == D("3000.000000")
    assert entry.unit_cost_at_movement == D("3000.000000")


def test_damage_requires_a_reason(db_session: Session, product: Product) -> None:
    """§12C.10: neden alanı zorunludur."""
    with pytest.raises(inventory.InventoryError, match="gerekçesiz"):
        inventory.damage(db_session, product=product, qty=D("1"), reason="   ")


def test_damage_report_shows_rate_and_cost(
    db_session: Session, store: Store, product: Product
) -> None:
    """§12C.10 raporu: fire gideri ve hasar oranı."""
    inventory.damage(db_session, product=product, qty=D("2"), reason="Kırıldı")
    inventory.record_movement(
        db_session,
        tenant_id=store.tenant_id,
        brand_id=store.brand_id,
        product_id=product.id,
        movement=InventoryMovement.SALE_OUT,
        qty=D("8"),
    )

    rows = inventory.damage_rows(db_session)

    assert len(rows) == 1
    row = rows[0]
    assert row.qty == D("2")
    assert row.cost == D("6000.0000")  # 2 × 3.000 ortalama
    assert row.damage_rate_pct == D("20.00")  # 2 / (2 + 8)


def test_damage_is_written_to_the_ledger_not_by_editing_history(
    db_session: Session, product: Product
) -> None:
    """Append-only: hasar geçmiş kaydı değiştirmez, yeni satır yazar."""
    before = len(db_session.scalars(select(InventoryLedger)).all())

    inventory.damage(db_session, product=product, qty=D("1"), reason="Nakliyede hasar")

    assert len(db_session.scalars(select(InventoryLedger)).all()) == before + 1


# --- MSRP ve marj tabanı (spec §12C.10) --------------------------------------


@pytest.fixture
def priced_product(db_session: Session, store: Store) -> Product:
    """MSRP'si olan, MSRP altında fiyatlanmış ürün."""
    record = make_product(db_session, store, "ALS-PSJS", cost=D("2980.0000"))
    record.msrp = D("6190.00")
    db_session.add(
        SkuPrice(
            product_id=record.id,
            store_id=store.id,
            price=D("5000.00"),
            effective_from=SALE_DATE,
            created_by="test",
        )
    )
    db_session.flush()
    return record


def test_price_below_msrp_is_a_violation(db_session: Session, priced_product: Product) -> None:
    """§12C.10: MSRP'nin ALTINDA satış marka disiplinini bozar."""
    rows = [
        row for row in discipline.violations(db_session, today=SALE_DATE) if "msrp" in row.kinds
    ]

    assert rows
    row = rows[0]
    assert row.sku == priced_product.sku
    assert row.msrp == D("6190.00")
    # (6190 − 5000) / 6190 = %19,22
    assert row.msrp_gap_pct == D("19.22")


def test_price_above_msrp_is_not_a_violation(
    db_session: Session, store: Store, priced_product: Product
) -> None:
    """MSRP üstü fiyat ihlal değildir — pazaryeri fiyatı serbesttir."""
    price = db_session.scalar(select(SkuPrice).where(SkuPrice.product_id == priced_product.id))
    assert price is not None
    price.price = D("6500.00")
    db_session.flush()

    rows = [
        row for row in discipline.violations(db_session, today=SALE_DATE) if "msrp" in row.kinds
    ]

    assert not [row for row in rows if row.sku == priced_product.sku]


def test_margin_floor_is_inherited_from_the_brand(
    db_session: Session, priced_product: Product
) -> None:
    """Ürün bazlı taban yoksa markanın varsayılanı geçerlidir (Alessi %18).

    Taban aktif markadan okunmalı: `brands` marka-kapsamlı bir tablo değildir, guard onu
    filtrelemez — yanlış markadan okunursa Alessi'nin tabanı Kahveji'ninkiyle ölçülürdü.
    """
    rows = discipline.violations(db_session, today=SALE_DATE)

    assert rows
    assert rows[0].floor_pct == D("18.00")
    # Marj (%23,5) tabanın üstünde: yalnızca MSRP ihlali var.
    assert rows[0].kinds == ("msrp",)


def test_margin_below_the_floor_is_a_violation(
    db_session: Session, priced_product: Product
) -> None:
    """Ürün bazlı taban markanınkini ezer; marj altına düşerse uyarı üretilir."""
    priced_product.min_margin_floor_pct = D("30.00")
    db_session.flush()

    rows = [
        row
        for row in discipline.violations(db_session, today=SALE_DATE)
        if "margin_floor" in row.kinds and row.sku == priced_product.sku
    ]

    assert rows
    assert rows[0].floor_pct == D("30.00")
    assert rows[0].margin_pct < D("30.00")


def test_alerts_are_written_once_per_day(db_session: Session, priced_product: Product) -> None:
    """Aynı ihlal için gün içinde ikinci uyarı yazılmaz — uyarı gürültüsü olmaz."""
    first = discipline.raise_alerts(db_session, today=SALE_DATE)
    second = discipline.raise_alerts(db_session, today=SALE_DATE)

    assert first > 0
    assert second == 0
    types = {alert.type for alert in db_session.scalars(select(Alert)).all()}
    assert discipline.MSRP_ALERT in types


# --- API ---------------------------------------------------------------------


@pytest.fixture
def api(db_session: Session, store: Store) -> Iterator[TestClient]:
    """Test oturumuna bağlı API istemcisi."""

    async def session_override() -> Any:
        yield db_session

    app = create_app()
    app.dependency_overrides[deps.get_session] = session_override
    with TestClient(app, raise_server_exceptions=False) as client:
        yield client


def _headers(api: TestClient, brand: str = "alessi") -> dict[str, str]:
    response = api.post("/auth/dev-login", json={"email": "mert@mokkalabs.com", "brand": brand})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def test_template_endpoint_returns_xlsx(api: TestClient, d2b: Store) -> None:
    """Şablon indirilebilir."""
    response = api.get("/alessi/b2b/template", headers=_headers(api))

    assert response.status_code == 200, response.text
    assert response.headers["content-type"] == api_b2b.TEMPLATE_MEDIA
    assert response.content[:2] == b"PK"  # xlsx = zip


def test_import_endpoint_previews_then_applies(
    api: TestClient, d2b: Store, product: Product
) -> None:
    """Önizleme yazmaz, onay yazar."""
    payload = _workbook([(SALE_DATE, "Kurumsal A.Ş.", "Bayi", product.sku, 2, 8000, 0, 20)])
    files = {"file": ("d2b.xlsx", payload)}

    preview = api.post(
        "/alessi/b2b/import", params={"dry_run": True}, files=files, headers=_headers(api)
    )
    assert preview.status_code == 200, preview.text
    assert preview.json()["orders"] == 0

    applied = api.post(
        "/alessi/b2b/import",
        params={"dry_run": False},
        files={"file": ("d2b.xlsx", payload)},
        headers=_headers(api),
    )
    assert applied.status_code == 200, applied.text
    assert applied.json()["orders"] == 1


def test_damage_endpoint_records_movement(api: TestClient, product: Product) -> None:
    """Hasar ucu deftere yazar."""
    response = api.post(
        "/alessi/inventory/damage",
        json={"product_id": str(product.id), "qty": "2", "reason": "Kırıldı — depo rafı"},
        headers=_headers(api),
    )

    assert response.status_code == 201, response.text
    assert response.json()["movement"] == "damage"


def test_damage_endpoint_requires_reason(api: TestClient, product: Product) -> None:
    """Gerekçesiz hasar kaydı reddedilir (şema seviyesinde)."""
    response = api.post(
        "/alessi/inventory/damage",
        json={"product_id": str(product.id), "qty": "2", "reason": "x"},
        headers=_headers(api),
    )

    assert response.status_code == 422


def test_discipline_endpoint_lists_violations(api: TestClient, priced_product: Product) -> None:
    """Disiplin ucu ihlalleri listeler."""
    response = api.get(
        "/alessi/discipline", params={"today": str(SALE_DATE)}, headers=_headers(api)
    )

    assert response.status_code == 200, response.text
    assert any(row["sku"] == priced_product.sku for row in response.json())


def test_modules_return_404_when_flags_are_off(api: TestClient, product: Product) -> None:
    """CLAUDE.md §2: kapalı modül 404 döner (403 değil)."""
    headers = _headers(api, brand="kahveji")

    assert api.get("/kahveji/b2b/template", headers=headers).status_code == 404
    assert api.get("/kahveji/b2b/tiers", headers=headers).status_code == 404
    assert api.get("/kahveji/discipline", headers=headers).status_code == 404


def test_customers_stay_inside_the_brand(db_session: Session, d2b: Store, product: Product) -> None:
    """CLAUDE.md §2: D2B müşterisi markaya yazılır, markalar arası sızmaz."""
    payload = _workbook([(SALE_DATE, "Kurumsal A.Ş.", "Bayi", product.sku, 1, 8000, 0, 20)])
    b2b.import_sales(db_session, payload=payload, store=d2b, dry_run=False)

    customer = db_session.scalar(select(Customer))
    assert customer is not None
    assert customer.brand_id == d2b.brand_id


def test_d2b_store_is_the_manual_channel(db_session: Session, d2b: Store) -> None:
    """D2B mağazası manuel kanaldadır (spec §12C.9)."""
    channel = db_session.scalar(select(Channel).where(Channel.id == d2b.channel_id))
    assert channel is not None
    assert channel.code is ChannelCode.MANUAL


def test_cancelled_lines_are_excluded_from_tier_summary(
    db_session: Session, d2b: Store, product: Product
) -> None:
    """İptal satır kademe özetine girmez."""
    payload = _workbook([(SALE_DATE, "Kurumsal A.Ş.", "Bayi", product.sku, 4, 8000, 0, 20)])
    b2b.import_sales(db_session, payload=payload, store=d2b, dry_run=False)
    line = db_session.scalar(select(OrderLine))
    assert line is not None
    line.status = OrderStatus.CANCELLED
    db_session.flush()

    rows = {row.tier: row for row in b2b.tier_margins(db_session)}

    assert rows["Bayi"].revenue == D("0.0000")


# --- satır doğrulama (hatalı satır diğerlerini engellemez) --------------------


@pytest.mark.parametrize(
    ("row", "expected"),
    [
        pytest.param((SALE_DATE, "", "bayi", None, 1, 100, 0, 20), "Müşteri boş", id="musteri-yok"),
        pytest.param(
            (SALE_DATE, "A", "bayi", None, 0, 100, 0, 20), "Adet pozitif", id="adet-sifir"
        ),
        pytest.param((SALE_DATE, "A", "bayi", None, 1, 0, 0, 20), "Birim fiyat", id="fiyat-sifir"),
        pytest.param(
            (SALE_DATE, "A", "bayi", None, 1, 100, 120, 20), "İskonto", id="iskonto-asiri"
        ),
        pytest.param(
            ("dün", "A", "bayi", None, 1, 100, 0, 20), "Tarih okunamadı", id="tarih-bozuk"
        ),
        pytest.param(
            (SALE_DATE, "A", "bayi", None, "abc", 100, 0, 20), "sayı değil", id="adet-metin"
        ),
    ],
)
def test_invalid_rows_are_rejected_with_reason(
    db_session: Session, d2b: Store, product: Product, row: tuple[Any, ...], expected: str
) -> None:
    """Her hatalı satır gerekçesiyle raporlanır; sessizce yok sayılmaz."""
    values = list(row)
    values[3] = product.sku
    payload = _workbook([tuple(values)])

    summary = b2b.import_sales(db_session, payload=payload, store=d2b, dry_run=False)

    assert summary.lines == 0
    assert expected in summary.errors[0].reason


def test_text_date_and_comma_decimal_are_accepted(
    db_session: Session, d2b: Store, product: Product
) -> None:
    """Kullanıcı Excel'de metin tarih ve virgüllü sayı yazabilir."""
    payload = _workbook(
        [("14.08.2026", "Kurumsal A.Ş.", "bayi", product.sku, "2", "8.000,50", "0", 20)]
    )

    summary = b2b.import_sales(db_session, payload=payload, store=d2b, dry_run=False)

    assert summary.errors == []
    line = db_session.scalar(select(OrderLine))
    assert line is not None
    assert line.unit_sale_price == D("8000.5000")


def test_customer_without_tier_falls_into_the_unknown_bucket(
    db_session: Session, d2b: Store, product: Product
) -> None:
    """Kademesi girilmemiş müşteri özet tablosunda "—" altında toplanır."""
    payload = _workbook([(SALE_DATE, "Kademesiz Ltd.", "", product.sku, 1, 8000, 0, 20)])
    b2b.import_sales(db_session, payload=payload, store=d2b, dry_run=False)

    rows = {row.tier: row for row in b2b.tier_margins(db_session)}

    assert "—" in rows
    assert rows["—"].revenue == D("8000.0000")


def test_foreign_brand_sku_is_rejected_in_d2b_import(
    db_session: Session, d2b: Store, product: Product
) -> None:
    """§3A.2: başka markanın SKU'su `cross_brand_rejected` ile reddedilir."""
    from app.core.context import system_scope
    from app.models.identity import Brand

    with system_scope():
        other = db_session.scalars(select(Brand).where(Brand.id != d2b.brand_id)).first()
        assert other is not None
        db_session.add(
            Product(
                tenant_id=other.tenant_id,
                brand_id=other.id,
                sku="KHV-YABANCI-1",
                name="Diğer markanın ürünü",
                vat_rate=D("1.00"),
            )
        )
        db_session.flush()

    payload = _workbook([(SALE_DATE, "Kurumsal A.Ş.", "bayi", "KHV-YABANCI-1", 1, 100, 0, 20)])
    summary = b2b.import_sales(db_session, payload=payload, store=d2b, dry_run=False)

    assert summary.lines == 0
    assert "cross_brand_rejected" in summary.errors[0].reason
