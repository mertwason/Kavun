"""KVN-12: senaryo motoru, karşılaştırma ve hedef marj çözücü (spec §12A.4).

Kabul kriteri (§12A.6): "hedef marj çözücüsünün sonucu, o fiyatla motor hesabına geri
verildiğinde hedef marjı ±0,01 puan tutturur" — round-trip testi + property testi.
"""

from __future__ import annotations

from collections.abc import Iterator
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

import pytest
from fastapi.testclient import TestClient
from hypothesis import HealthCheck, given, settings
from hypothesis import strategies as st
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api import deps
from app.core.context import RequestContext, system_scope, use_context
from app.engine.pricing import PriceInputs, break_even_price, price_for_margin
from app.engine.profit import LineInput, compute_line_profit
from app.main import create_app
from app.models.catalog import Product
from app.models.enums import CommissionMode, ShippingPayer, UserRole
from app.models.identity import Brand, Store
from app.models.workspace import PricingScenario
from app.services import scenarios
from tests.profit_factories import make_commission, make_product, make_store

D = Decimal
TODAY = date(2026, 8, 19)
CATEGORY = "Kahve/Harman"


@pytest.fixture
def store(db_session: Session) -> Iterator[Store]:
    """Mağaza + tarife + hizmet bedeli; marka bağlamı kurulur."""
    with system_scope():
        store = make_store(db_session)
        make_commission(db_session, store, rate=D("0.2000"), category=CATEGORY)
        store.service_fee_per_order = D("12.0000")
        db_session.flush()
        brand = db_session.get(Brand, store.brand_id)
    assert brand is not None
    context = RequestContext(
        tenant_id=brand.tenant_id,
        user_id=None,
        brand_id=brand.id,
        brand_slug=brand.slug,
        role=UserRole.ADMIN,
    )
    with use_context(context):
        yield store


@pytest.fixture
def product(db_session: Session, store: Store) -> Product:
    """50 TL net maliyetli, %20 KDV'li ürün."""
    return make_product(db_session, store, "SENARYO-1", cost=D("50.0000"), category=CATEGORY)


def _scenario(**overrides: Any) -> scenarios.ScenarioInput:
    values: dict[str, Any] = {
        "name": "Baz",
        "satis_fiyati": D("199.00"),
        "kargo_tahmini": D("24.00"),
    }
    values.update(overrides)
    return scenarios.ScenarioInput(**values)


# --- çözücü: cebirsel doğruluk (spec §12A.4, kabul kriteri §12A.6) -----------


def _engine_margin(price: Decimal, inputs: PriceInputs) -> Decimal:
    """Çözülen fiyatı motora geri verir ve marjı okur."""
    discount = price * inputs.campaign_discount_rate
    result = compute_line_profit(
        LineInput(
            line_gross=price - discount,
            qty=1,
            vat_percent=inputs.vat_percent,
            unit_cost_net=inputs.unit_cost_net,
            commission_rate=inputs.commission_rate,
            cargo_cost=inputs.cargo_cost,
            service_fee=inputs.service_fee,
            campaign_discount=discount,
            campaign_seller_share_rate=inputs.campaign_seller_share_rate,
            service_vat_percent=inputs.service_vat_percent,
        )
    )
    return result.margin_pct


BASE_INPUTS = PriceInputs(
    unit_cost_net=D("50"),
    vat_percent=D("20"),
    commission_rate=D("0.20"),
    cargo_cost=D("24"),
    service_fee=D("12"),
)


@pytest.mark.parametrize("target", [D("0"), D("10"), D("20"), D("35"), D("-5")])
def test_solved_price_hits_the_target_margin(target: Decimal) -> None:
    """§12A.6 kabul kriteri: çözülen fiyat motora geri verilince hedef marj ±0,01 tutar."""
    price = price_for_margin(target, BASE_INPUTS)

    assert price is not None
    assert abs(_engine_margin(price, BASE_INPUTS) - target) <= D("0.01")


def test_solver_handles_campaign_discount() -> None:
    """Kampanya indirimi ve platform payı varken de hedef marj tutar."""
    inputs = PriceInputs(
        unit_cost_net=D("60"),
        vat_percent=D("20"),
        commission_rate=D("0.20"),
        cargo_cost=D("30"),
        service_fee=D("10"),
        campaign_discount_rate=D("0.10"),
        campaign_seller_share_rate=D("0.5"),
    )

    price = price_for_margin(D("15"), inputs)

    assert price is not None
    assert abs(_engine_margin(price, inputs) - D("15")) <= D("0.01")


def test_break_even_price_matches_engine_zero_profit() -> None:
    """Başabaş fiyat motorda tam sıfır kâr verir (KVN-07'nin elle hesabıyla da uyumlu)."""
    price = break_even_price(BASE_INPUTS)

    assert price == D("120.0000")
    assert _engine_margin(price, BASE_INPUTS) == D("0.0000")


def test_unreachable_target_returns_none_instead_of_fantasy_price() -> None:
    """Komisyon + KDV yapısıyla ulaşılamayan marj için fiyat UYDURULMAZ."""
    assert price_for_margin(D("90"), BASE_INPUTS) is None


def test_costless_product_has_no_meaningful_target_price() -> None:
    """Maliyet ve gider yoksa her fiyat hedefi sağlar; anlamlı bir fiyat YOK → None."""
    free = PriceInputs(
        unit_cost_net=D("0"),
        vat_percent=D("20"),
        commission_rate=D("0.20"),
        cargo_cost=D("0"),
        service_fee=D("0"),
    )

    assert price_for_margin(D("10"), free) is None


@given(
    cost=st.decimals(min_value=D("1"), max_value=D("5000"), places=2),
    commission=st.decimals(min_value=D("0"), max_value=D("0.35"), places=4),
    cargo=st.decimals(min_value=D("0"), max_value=D("500"), places=2),
    service=st.decimals(min_value=D("0"), max_value=D("100"), places=2),
    vat=st.sampled_from([D("0"), D("1"), D("10"), D("20")]),
    target=st.decimals(min_value=D("-20"), max_value=D("40"), places=2),
)
@settings(
    max_examples=200, deadline=None, suppress_health_check=[HealthCheck.function_scoped_fixture]
)
def test_property_solver_round_trip(
    cost: Decimal,
    commission: Decimal,
    cargo: Decimal,
    service: Decimal,
    vat: Decimal,
    target: Decimal,
) -> None:
    """Değişmez: çözücü bir fiyat döndürdüyse o fiyat hedef marjı ±0,01 tutturur."""
    inputs = PriceInputs(
        unit_cost_net=cost,
        vat_percent=vat,
        commission_rate=commission,
        cargo_cost=cargo,
        service_fee=service,
    )
    price = price_for_margin(target, inputs)
    if price is None:
        return
    assert abs(_engine_margin(price, inputs) - target) <= D("0.01")


# --- senaryo hesabı (spec §12A.4) -------------------------------------------


def test_scenario_result_matches_engine(
    db_session: Session, store: Store, product: Product
) -> None:
    """Senaryo sonucu motorun çıktısıdır — ikinci bir formül yok."""
    result = scenarios.evaluate(
        db_session, product=product, store=store, scenario=_scenario(), on_date=TODAY
    )

    expected = compute_line_profit(
        LineInput(
            line_gross=D("199.00"),
            qty=1,
            vat_percent=product.vat_rate,
            unit_cost_net=D("50.0000"),
            commission_rate=D("0.2000"),
            cargo_cost=D("24.00"),
            service_fee=D("12.0000"),
        )
    )
    assert result.birim_kar == expected.profit
    assert result.marj_pct == expected.margin_pct


def test_total_profit_scales_with_quantity_assumption(
    db_session: Session, store: Store, product: Product
) -> None:
    """§12A.4: toplam kâr = birim kâr × adet varsayımı (talep TAHMİNİ değil, varsayım)."""
    single = scenarios.evaluate(
        db_session, product=product, store=store, scenario=_scenario(), on_date=TODAY
    )
    hundred = scenarios.evaluate(
        db_session,
        product=product,
        store=store,
        scenario=_scenario(adet_varsayimi=100),
        on_date=TODAY,
    )

    assert hundred.toplam_kar == single.birim_kar * 100
    assert hundred.birim_kar == single.birim_kar


def test_buyer_paid_shipping_removes_seller_cargo_cost(
    db_session: Session, store: Store, product: Product
) -> None:
    """§12A.4: kargoyu alıcı öderse satıcının kargo maliyeti sıfırlanır."""
    seller = scenarios.evaluate(
        db_session, product=product, store=store, scenario=_scenario(), on_date=TODAY
    )
    buyer = scenarios.evaluate(
        db_session,
        product=product,
        store=store,
        scenario=_scenario(kargo_kim_oder=ShippingPayer.ALICI),
        on_date=TODAY,
    )

    assert buyer.cargo_cost == D("0")
    assert buyer.birim_kar > seller.birim_kar


def test_campaign_discount_reduces_customer_price(
    db_session: Session, store: Store, product: Product
) -> None:
    """§12A.4: indirim müşterinin ödediği tutarı düşürür; liste fiyatı korunur."""
    result = scenarios.evaluate(
        db_session,
        product=product,
        store=store,
        scenario=_scenario(kampanya_indirim_pct=D("10"), kampanya_satici_pay_pct=D("100")),
        on_date=TODAY,
    )

    assert result.satis_fiyati == D("199.0000")
    assert result.musteri_odedigi == D("179.1000")


def test_platform_funded_campaign_beats_seller_funded(
    db_session: Session, store: Store, product: Product
) -> None:
    """§12A.4: indirimin platform payı arttıkça satıcının kârı artar."""
    seller_pays = scenarios.evaluate(
        db_session,
        product=product,
        store=store,
        scenario=_scenario(kampanya_indirim_pct=D("20"), kampanya_satici_pay_pct=D("100")),
        on_date=TODAY,
    )
    shared = scenarios.evaluate(
        db_session,
        product=product,
        store=store,
        scenario=_scenario(kampanya_indirim_pct=D("20"), kampanya_satici_pay_pct=D("50")),
        on_date=TODAY,
    )

    assert shared.birim_kar > seller_pays.birim_kar


def test_pinned_commission_overrides_tariff(
    db_session: Session, store: Store, product: Product
) -> None:
    """§12B.4: senaryoda oran sabitlenebilir; tarife değişse de senaryo aynı kalır."""
    result = scenarios.evaluate(
        db_session,
        product=product,
        store=store,
        scenario=_scenario(
            commission_mode=CommissionMode.PINNED, pinned_commission_rate=D("0.3000")
        ),
        on_date=TODAY,
    )

    assert result.commission_rate == D("0.3000")
    assert result.breakdown.cost_commission == D("59.7000")  # 199 × %30


def test_break_even_is_reported_per_scenario(
    db_session: Session, store: Store, product: Product
) -> None:
    """§12A.4: her senaryo başabaş fiyatını da döner."""
    result = scenarios.evaluate(
        db_session, product=product, store=store, scenario=_scenario(), on_date=TODAY
    )

    assert result.basabas_fiyat == D("120.0000")


def test_compare_rejects_more_than_three(
    db_session: Session, store: Store, product: Product
) -> None:
    """§12A.4: karşılaştırmada en fazla 3 senaryo."""
    items = [(product, store, _scenario(name=f"S{index}")) for index in range(4)]

    with pytest.raises(scenarios.ScenarioError, match=r"[eE]n fazla 3"):
        scenarios.compare(db_session, scenarios=items, on_date=TODAY)


def test_compare_returns_results_in_order(
    db_session: Session, store: Store, product: Product
) -> None:
    """Karşılaştırma sonuçları verilen sırayı korur (UI sütunları kaymasın)."""
    items = [
        (product, store, _scenario(name="Ucuz", satis_fiyati=D("149.00"))),
        (product, store, _scenario(name="Orta", satis_fiyati=D("199.00"))),
        (product, store, _scenario(name="Pahalı", satis_fiyati=D("249.00"))),
    ]

    results = scenarios.compare(db_session, scenarios=items, on_date=TODAY)

    assert [result.name for result in results] == ["Ucuz", "Orta", "Pahalı"]
    assert results[0].birim_kar < results[1].birim_kar < results[2].birim_kar


# --- hedef marj çözücü, DB katmanında ---------------------------------------


def test_solve_target_margin_returns_price_that_hits_the_target(
    db_session: Session, store: Store, product: Product
) -> None:
    """§12A.6: çözülen fiyatla yapılan senaryo hesabı hedef marjı ±0,01 tutturur."""
    solved = scenarios.solve_target_margin(
        db_session,
        product=product,
        store=store,
        target_margin_pct=D("25"),
        scenario=_scenario(),
        on_date=TODAY,
    )

    assert solved.reachable
    assert solved.result is not None
    assert abs(solved.result.marj_pct - D("25")) <= D("0.01")


def test_solve_without_cost_is_honest(db_session: Session, store: Store) -> None:
    """Maliyet yoksa fiyat uydurulmaz; sebep yazılır."""
    product = make_product(db_session, store, "MALIYETSIZ-1", category=CATEGORY)
    with system_scope():
        from app.models.catalog import SkuCost

        for cost in db_session.scalars(
            select(SkuCost).where(SkuCost.product_id == product.id)
        ).all():
            db_session.delete(cost)
        db_session.flush()

    solved = scenarios.solve_target_margin(
        db_session,
        product=product,
        store=store,
        target_margin_pct=D("25"),
        scenario=_scenario(),
        on_date=TODAY,
    )

    assert not solved.reachable
    assert solved.price is None
    assert "maliyeti tanımlı değil" in solved.message


# --- API katmanı -------------------------------------------------------------


@pytest.fixture
def api(db_session: Session, store: Store, product: Product) -> Iterator[TestClient]:
    """Test oturumuna bağlı API istemcisi."""

    async def session_override() -> Any:
        yield db_session

    app = create_app()
    app.dependency_overrides[deps.get_session] = session_override
    with TestClient(app, raise_server_exceptions=False) as client:
        yield client


def _headers(api: TestClient, brand: str = "alessi") -> dict[str, str]:
    response = api.post("/auth/dev-login", json={"email": "mert@mokkalabs.com", "brand": brand})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def _body(product: Product, **overrides: Any) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "name": "API Senaryosu",
        "product_id": str(product.id),
        "satis_fiyati": "199.00",
        "kargo_tahmini": "24.00",
    }
    payload.update(overrides)
    return payload


def test_evaluate_endpoint_writes_nothing(
    api: TestClient, db_session: Session, product: Product
) -> None:
    """Hesap ucu kayıt oluşturmaz."""
    response = api.post("/alessi/scenarios/evaluate", json=_body(product), headers=_headers(api))

    assert response.status_code == 200, response.text
    assert db_session.scalar(select(PricingScenario)) is None


def test_create_and_list_scenarios(api: TestClient, product: Product) -> None:
    """Senaryo kaydedilir ve listede güncel hesapla döner."""
    created = api.post("/alessi/scenarios", json=_body(product), headers=_headers(api))
    assert created.status_code == 201, created.text

    listed = api.get("/alessi/scenarios", headers=_headers(api))

    assert listed.status_code == 200
    assert [item["name"] for item in listed.json()] == ["API Senaryosu"]
    assert listed.json()[0]["basabas_fiyat"] == "120.0000"


def test_compare_endpoint_rejects_four(api: TestClient, product: Product) -> None:
    """§12A.4: 4 senaryo 422 döner."""
    response = api.post(
        "/alessi/scenarios/compare",
        json={"inputs": [_body(product, name=f"S{index}") for index in range(4)]},
        headers=_headers(api),
    )

    assert response.status_code == 422


def test_target_margin_endpoint_round_trip(api: TestClient, product: Product) -> None:
    """§12A.6: uç, hedef marjı tutturan fiyatı ve o fiyatın sonucunu döner."""
    response = api.post(
        "/alessi/scenarios/target-margin",
        json=_body(product) | {"hedef_marj_pct": "25"},
        headers=_headers(api),
    )

    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["reachable"] is True
    assert abs(Decimal(payload["result"]["marj_pct"]) - D("25")) <= D("0.01")


def test_target_margin_unreachable_is_reported(api: TestClient, product: Product) -> None:
    """Ulaşılamayan hedef için fiyat yerine açıklama döner."""
    response = api.post(
        "/alessi/scenarios/target-margin",
        json=_body(product) | {"hedef_marj_pct": "95"},
        headers=_headers(api),
    )

    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["reachable"] is False
    assert payload["price"] is None
    assert "ulaşılamaz" in payload["message"]


def test_scenario_of_other_brand_product_returns_404(api: TestClient, product: Product) -> None:
    """§3A.6: başka markanın ürünü için senaryo kurulamaz."""
    response = api.post(
        "/kahveji/scenarios/evaluate", json=_body(product), headers=_headers(api, "kahveji")
    )

    assert response.status_code == 404


# --- senaryo xlsx round-trip (spec §12A.4) ----------------------------------


def test_scenario_workbook_round_trip(db_session: Session, store: Store, product: Product) -> None:
    """§12A.4: export → import → hesaplanmış sonuç sütunlarıyla dolu dosya döner."""
    inputs = [_scenario(name="Ucuz", satis_fiyati=D("149.00")), _scenario(name="Pahalı")]
    results = scenarios.compare(
        db_session, scenarios=[(product, store, item) for item in inputs], on_date=TODAY
    )
    payload = scenarios.build_scenario_workbook(results, inputs=inputs)

    parsed = scenarios.parse_scenario_workbook(db_session, payload)

    assert [scenario.name for _, scenario in parsed] == ["Ucuz", "Pahalı"]
    assert [scenario.satis_fiyati for _, scenario in parsed] == [D("149.0000"), D("199.0000")]
    assert [scenario.kargo_tahmini for _, scenario in parsed] == [D("24.00"), D("24.00")]

    recomputed = scenarios.compare(
        db_session, scenarios=[(product, store, item) for _, item in parsed], on_date=TODAY
    )
    assert [item.birim_kar for item in recomputed] == [item.birim_kar for item in results]


def test_scenario_workbook_computed_columns_are_ignored(
    db_session: Session, store: Store, product: Product
) -> None:
    """Kullanıcı `Birim Kâr` sütununu elle değiştirse bile hesap motordan gelir."""
    inputs = [_scenario()]
    results = scenarios.compare(db_session, scenarios=[(product, store, inputs[0])], on_date=TODAY)
    payload = scenarios.build_scenario_workbook(results, inputs=inputs)

    workbook = load_workbook(BytesIO(payload))
    sheet = workbook[scenarios.SCENARIO_SHEET]
    headers = [cell.value for cell in sheet[2]]
    sheet.cell(row=3, column=headers.index("Birim Kâr") + 1).value = 999999
    buffer = BytesIO()
    workbook.save(buffer)

    parsed = scenarios.parse_scenario_workbook(db_session, buffer.getvalue())
    recomputed = scenarios.compare(
        db_session, scenarios=[(product, store, parsed[0][1])], on_date=TODAY
    )

    assert recomputed[0].birim_kar == results[0].birim_kar


def test_scenario_workbook_rejects_unknown_sku(
    db_session: Session, store: Store, product: Product
) -> None:
    """Bilinmeyen SKU sessizce atlanmaz; dosya reddedilir."""
    inputs = [_scenario()]
    results = scenarios.compare(db_session, scenarios=[(product, store, inputs[0])], on_date=TODAY)
    payload = scenarios.build_scenario_workbook(results, inputs=inputs)

    workbook = load_workbook(BytesIO(payload))
    workbook[scenarios.SCENARIO_SHEET].cell(row=3, column=1).value = "YOK-1"
    buffer = BytesIO()
    workbook.save(buffer)

    with pytest.raises(scenarios.ScenarioError, match="bilinmeyen SKU"):
        scenarios.parse_scenario_workbook(db_session, buffer.getvalue())


def test_scenario_workbook_rejects_foreign_template(db_session: Session) -> None:
    """Şablon sürümü tutmayan dosya işlenmez."""
    with pytest.raises(scenarios.ScenarioError, match=r"okunamadı|Şablon"):
        scenarios.parse_scenario_workbook(db_session, b"excel degil")


def test_scenario_export_import_endpoints(api: TestClient, product: Product) -> None:
    """Uçtan uca: senaryo kaydet → export → import → hesaplanmış dosya."""
    api.post("/alessi/scenarios", json=_body(product), headers=_headers(api))

    exported = api.get("/alessi/scenarios/export", headers=_headers(api))
    assert exported.status_code == 200, exported.text
    assert scenarios.SCENARIO_SHEET in load_workbook(BytesIO(exported.content)).sheetnames

    response = api.post(
        "/alessi/scenarios/import",
        files={"file": ("senaryolar.xlsx", exported.content)},
        headers=_headers(api),
    )

    assert response.status_code == 200, response.text
    sheet = load_workbook(BytesIO(response.content))[scenarios.SCENARIO_SHEET]
    headers = [cell.value for cell in sheet[2]]
    assert sheet.cell(row=3, column=headers.index("Birim Kâr") + 1).value is not None
