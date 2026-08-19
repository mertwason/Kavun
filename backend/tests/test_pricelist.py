"""KVN-10: fiyat listesi Excel round-trip'i (spec §12A.1, §12A.2, §12A.6).

Kabul kriterleri (§12A.6) birebir test edilir:
- export → değiştirmeden import → dry_run: 0 yeni, 0 güncelleme, 0 hata
- 500 satırlık dosya < 10 sn
"""

from __future__ import annotations

import time
from collections.abc import Iterator
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.api import deps
from app.core.context import RequestContext, system_scope, use_context
from app.main import create_app
from app.models.catalog import Product, SkuCost, SkuLogistics, SkuPrice
from app.models.enums import UserRole
from app.models.identity import Brand, Channel, Store
from app.models.workspace import ImportBatch
from app.services import pricelist
from tests.profit_factories import make_commission, make_product, make_store

D = Decimal
TODAY = date(2026, 8, 19)


@pytest.fixture
def store(db_session: Session) -> Iterator[Store]:
    """Alessi Trendyol mağazası + iki ürün + kategori tarifesi.

    Fixture marka BAĞLAMI kurar (system_scope değil): fiyat listesi gerçek hayatta
    workspace içinde üretilir ve guard'ın markayı kısıtladığını test de görmelidir.
    Alessi seçilir çünkü İKİ kanalı var (trendyol + manual/D2B) — aynı SKU'nun birden
    fazla satırı olduğu, dolayısıyla daha zor olan durum.
    """
    with system_scope():
        store = make_store(db_session)
        make_commission(db_session, store)
        for sku in ("KHV-A", "KHV-B"):
            product = make_product(db_session, store, sku)
            db_session.add(
                SkuPrice(
                    product_id=product.id,
                    store_id=store.id,
                    price=D("120.0000"),
                    effective_from=TODAY - timedelta(days=30),
                )
            )
        db_session.flush()
        brand = db_session.get(Brand, store.brand_id)
    assert brand is not None
    context = RequestContext(
        tenant_id=brand.tenant_id,
        user_id=None,
        brand_id=brand.id,
        brand_slug=brand.slug,
        role=UserRole.ADMIN,
    )
    with use_context(context):
        yield store


def _export(db_session: Session, store: Store) -> bytes:
    brand = db_session.get(Brand, store.brand_id)
    assert brand is not None
    return pricelist.export_price_list(db_session, brand_name=brand.name, today=TODAY)


def _edit(payload: bytes, row_no: int, column: str, value: Any) -> bytes:
    """Dosyadaki bir hücreyi değiştirir (kullanıcının Excel'de yaptığı şey)."""
    workbook = load_workbook(BytesIO(payload))
    sheet = workbook[pricelist.SHEET_NAME]
    headers = [cell.value for cell in sheet[pricelist.HEADER_ROW]]
    # `ws.cell(..., value=None)` hücreyi TEMİZLEMEZ (openpyxl "değer verilmedi" sayar);
    # boş hücre senaryosunu test edebilmek için doğrudan atanır.
    sheet.cell(row=row_no, column=headers.index(column) + 1).value = value
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _edit_all_rows_of(payload: bytes, sku: str, column: str, value: Any) -> bytes:
    """Bir SKU'nun TÜM kanal satırlarında aynı hücreyi değiştirir.

    Maliyet ve desi ürün seviyesindedir; yalnızca bir kanal satırını değiştirmek
    dosyayı kendi içinde çelişkiye düşürür (import bunu hata sayar).
    """
    workbook = load_workbook(BytesIO(payload))
    sheet = workbook[pricelist.SHEET_NAME]
    headers = [cell.value for cell in sheet[pricelist.HEADER_ROW]]
    sku_column = headers.index("SKU") + 1
    target = headers.index(column) + 1
    for row_no in range(pricelist.FIRST_DATA_ROW, sheet.max_row + 1):
        if sheet.cell(row=row_no, column=sku_column).value == sku:
            sheet.cell(row=row_no, column=target).value = value
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _channel_code(db_session: Session, store: Store) -> str:
    channel = db_session.get(Channel, store.channel_id)
    assert channel is not None
    return str(channel.code.value)


def _rows(payload: bytes) -> list[tuple[Any, ...]]:
    sheet = load_workbook(BytesIO(payload))[pricelist.SHEET_NAME]
    return list(sheet.iter_rows(min_row=pricelist.FIRST_DATA_ROW, values_only=True))


# --- export (spec §12A.1) ----------------------------------------------------


def test_export_carries_template_version_and_columns(db_session: Session, store: Store) -> None:
    """§12A.1: ilk satırda şablon sürümü, ikinci satırda sabit sütun başlıkları."""
    sheet = load_workbook(BytesIO(_export(db_session, store)))[pricelist.SHEET_NAME]

    assert sheet[pricelist.META_CELL].value == pricelist.TEMPLATE_VERSION
    headers = [cell.value for cell in sheet[pricelist.HEADER_ROW]]
    assert headers == list(pricelist.COLUMNS)


def test_export_profit_columns_come_from_the_engine(db_session: Session, store: Store) -> None:
    """§12A.1: `Net Kâr` motorun hesabıdır — Excel'de ikinci bir formül yaşamaz."""
    columns = list(pricelist.COLUMNS)
    channel = _channel_code(db_session, store)
    row = next(
        item
        for item in _rows(_export(db_session, store))
        if item[columns.index("Kanal")] == channel
    )

    product = db_session.scalar(select(Product).where(Product.sku == row[columns.index("SKU")]))
    assert product is not None
    expected_profit, expected_margin, _ = pricelist.preview_profit(
        db_session,
        product,
        store,
        price=D(str(row[columns.index("Satış Fiyatı")])),
        unit_cost=D(str(row[columns.index("Alış Maliyeti")])),
        cargo=D("0"),
        on_date=TODAY,
    )
    assert D(str(row[columns.index("Net Kâr")])) == expected_profit
    assert D(str(row[columns.index("Marj %")])) == expected_margin


# --- kabul kriteri: round-trip idempotency (spec §12A.6) ---------------------


def test_unchanged_roundtrip_reports_no_changes(db_session: Session, store: Store) -> None:
    """§12A.6 kabul kriteri: export → değiştirmeden import → 0 yeni, 0 güncelleme, 0 hata."""
    summary = pricelist.import_price_list(
        db_session, _export(db_session, store), today=TODAY, dry_run=True
    )

    assert (summary.yeni, summary.guncelleme, summary.hata) == (0, 0, 0)
    assert summary.degisiklik_yok == len(summary.rows)


def test_computed_columns_are_ignored_on_import(db_session: Session, store: Store) -> None:
    """§12A.1: kullanıcı `Net Kâr` sütununu elle değiştirse bile import onu yok sayar."""
    edited = _edit(_export(db_session, store), pricelist.FIRST_DATA_ROW, "Net Kâr", 999999)

    summary = pricelist.import_price_list(db_session, edited, today=TODAY, dry_run=True)

    assert (summary.yeni, summary.guncelleme, summary.hata) == (0, 0, 0)


# --- diff ve uygulama (spec §12A.2) -----------------------------------------


def test_price_change_is_reported_as_update_with_diff(db_session: Session, store: Store) -> None:
    """§12A.2.1: değişiklik `guncelleme` olarak, eski → yeni diff'iyle raporlanır."""
    edited = _edit(_export(db_session, store), pricelist.FIRST_DATA_ROW, "Satış Fiyatı", 149.9)

    summary = pricelist.import_price_list(db_session, edited, today=TODAY, dry_run=True)

    assert summary.guncelleme == 1
    changed = next(row for row in summary.rows if row.action == "guncelleme")
    assert "fiyat" in changed.changes
    assert "149.9" in changed.changes["fiyat"]


def test_dry_run_writes_nothing(db_session: Session, store: Store) -> None:
    """§12A.2.1: `dry_run=true` iken tek bir satır bile yazılmaz."""
    before = db_session.scalar(select(func.count(SkuPrice.id)))
    edited = _edit(_export(db_session, store), pricelist.FIRST_DATA_ROW, "Satış Fiyatı", 149.9)

    pricelist.import_price_list(db_session, edited, today=TODAY, dry_run=True)

    assert db_session.scalar(select(func.count(SkuPrice.id))) == before


def test_apply_writes_versioned_records_and_keeps_history(
    db_session: Session, store: Store
) -> None:
    """§12A.1: maliyet/fiyat değişikliği ESKİ kaydı ezmez, yeni versiyon ekler."""
    exported = _export(db_session, store)
    sku = _rows(exported)[0][0]
    edited = _edit_all_rows_of(exported, sku, "Alış Maliyeti", 77.5)

    summary = pricelist.import_price_list(db_session, edited, today=TODAY, dry_run=False)

    assert summary.guncelleme >= 1
    product = db_session.scalar(select(Product).where(Product.sku == sku))
    assert product is not None
    costs = list(
        db_session.scalars(
            select(SkuCost).where(SkuCost.product_id == product.id).order_by(SkuCost.effective_from)
        ).all()
    )
    assert len(costs) == 2  # eski kayıt duruyor
    assert costs[-1].unit_cost == D("77.5")
    assert costs[-1].effective_from == TODAY


def test_second_import_of_the_same_file_is_idempotent(db_session: Session, store: Store) -> None:
    """Aynı dosya iki kez uygulanırsa ikincisi hiçbir şey değiştirmez."""
    edited = _edit(_export(db_session, store), pricelist.FIRST_DATA_ROW, "Satış Fiyatı", 149.9)

    pricelist.import_price_list(db_session, edited, today=TODAY, dry_run=False)
    second = pricelist.import_price_list(db_session, edited, today=TODAY, dry_run=False)

    assert (second.yeni, second.guncelleme, second.hata) == (0, 0, 0)


def test_new_sku_creates_product_cost_and_price(db_session: Session, store: Store) -> None:
    """§12A.1: yeni SKU satırı ürün + maliyet + fiyat kaydı doğurur."""
    payload = _export(db_session, store)
    workbook = load_workbook(BytesIO(payload))
    sheet = workbook[pricelist.SHEET_NAME]
    sheet.append(("KHV-YENI", "Yeni Ürün", "Kahveji", "trendyol", 20, 1.5, 60, 199, None, 0, 12))
    buffer = BytesIO()
    workbook.save(buffer)

    summary = pricelist.import_price_list(db_session, buffer.getvalue(), today=TODAY, dry_run=False)

    assert summary.yeni == 1
    product = db_session.scalar(select(Product).where(Product.sku == "KHV-YENI"))
    assert product is not None
    assert product.vat_rate == D("20")
    assert db_session.scalar(select(SkuCost).where(SkuCost.product_id == product.id)) is not None
    assert db_session.scalar(select(SkuPrice).where(SkuPrice.product_id == product.id)) is not None
    assert (
        db_session.scalar(select(SkuLogistics).where(SkuLogistics.product_id == product.id))
        is not None
    )


# --- doğrulama (spec §12A.2.3) ----------------------------------------------


@pytest.mark.parametrize(
    ("column", "value", "expected"),
    [
        ("SKU", None, "SKU boş"),
        ("Kanal", "hepsiburada", "Bilinmeyen kanal"),
        ("KDV %", 7, "Geçersiz KDV oranı"),
        ("KDV %", None, "KDV % zorunlu"),
        ("Satış Fiyatı", -10, "negatif olamaz"),
        ("Alış Maliyeti", "abc", "sayı değil"),
    ],
)
def test_invalid_rows_are_rejected_with_reason(
    db_session: Session, store: Store, column: str, value: Any, expected: str
) -> None:
    """§12A.2.3: hatalı satır reddedilir ve NEDEN reddedildiği yazılır."""
    edited = _edit(_export(db_session, store), pricelist.FIRST_DATA_ROW, column, value)

    summary = pricelist.import_price_list(db_session, edited, today=TODAY, dry_run=True)

    assert summary.hata == 1
    failed = next(row for row in summary.rows if row.action == "hata")
    assert expected in failed.message


def test_conflicting_product_level_value_is_rejected(db_session: Session, store: Store) -> None:
    """Aynı SKU'nun kanal satırlarında maliyet farklıysa dosya reddedilir.

    Maliyet ürünün özelliğidir; kullanıcı iki kanal satırından yalnızca birini
    düzenlerse dosya kendi içinde çelişir. "Son satır kazansın" demek sessiz veri
    kaybı olurdu — o SKU'nun tüm satırları hata olarak işaretlenir.
    """
    edited = _edit(_export(db_session, store), pricelist.FIRST_DATA_ROW, "Alış Maliyeti", 77.5)

    summary = pricelist.import_price_list(db_session, edited, today=TODAY, dry_run=True)

    assert summary.hata == 2  # çelişen SKU'nun her iki kanal satırı da reddedildi
    failed = next(row for row in summary.rows if row.action == "hata")
    assert "farklı Alış Maliyeti" in failed.message


def test_conflict_check_ignores_channel_level_price(db_session: Session, store: Store) -> None:
    """Fiyat kanal bazlıdır: kanaldan kanala farklı olması ÇELİŞKİ DEĞİLDİR."""
    edited = _edit(_export(db_session, store), pricelist.FIRST_DATA_ROW, "Satış Fiyatı", 199.9)

    summary = pricelist.import_price_list(db_session, edited, today=TODAY, dry_run=True)

    assert summary.hata == 0
    assert summary.guncelleme == 1


def test_invalid_row_does_not_block_valid_rows(db_session: Session, store: Store) -> None:
    """Bir satırın hatası diğerlerini düşürmez — dosyanın tamamı işlenir."""
    edited = _edit(_export(db_session, store), pricelist.FIRST_DATA_ROW, "KDV %", 7)

    summary = pricelist.import_price_list(db_session, edited, today=TODAY, dry_run=True)

    assert summary.hata == 1
    assert summary.degisiklik_yok == len(summary.rows) - 1


def test_foreign_template_is_rejected(db_session: Session, store: Store) -> None:
    """§12A.1: şablon sürümü tutmayan dosya işlenmez — sessizce yanlış veri yazılmaz."""
    payload = _export(db_session, store)
    workbook = load_workbook(BytesIO(payload))
    workbook[pricelist.SHEET_NAME][pricelist.META_CELL] = "baska-sablon-v9"
    buffer = BytesIO()
    workbook.save(buffer)

    with pytest.raises(pricelist.TemplateError, match="Şablon sürümü"):
        pricelist.import_price_list(db_session, buffer.getvalue(), today=TODAY)


def test_non_excel_payload_is_rejected(db_session: Session, store: Store) -> None:
    """Excel olmayan dosya anlaşılır hata verir."""
    with pytest.raises(pricelist.TemplateError, match="okunamadı"):
        pricelist.import_price_list(db_session, b"bu bir excel degil", today=TODAY)


def test_error_workbook_lists_failed_rows(db_session: Session, store: Store) -> None:
    """§12A.2.3: hatalı satırlar orijinal dosyaya "Hatalar" sayfası olarak eklenir."""
    edited = _edit(_export(db_session, store), pricelist.FIRST_DATA_ROW, "KDV %", 7)
    summary = pricelist.import_price_list(db_session, edited, today=TODAY, dry_run=True)

    report = pricelist.error_workbook(edited, summary)

    sheet = load_workbook(BytesIO(report))[pricelist.ERROR_SHEET_NAME]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][0] == pricelist.FIRST_DATA_ROW
    assert "KDV" in str(rows[0][3])


# --- kabul kriteri: performans (spec §12A.6) --------------------------------


def test_five_hundred_rows_process_under_ten_seconds(db_session: Session, store: Store) -> None:
    """§12A.6 kabul kriteri: 500 satırlık dosya 10 saniyenin altında işlenir."""
    payload = _export(db_session, store)
    workbook = load_workbook(BytesIO(payload))
    sheet = workbook[pricelist.SHEET_NAME]
    for index in range(500):
        sheet.append(
            (
                f"PERF-{index:04d}",
                f"Perf {index}",
                "Kahveji",
                "trendyol",
                20,
                1,
                50,
                120,
                None,
                0,
                12,
            )
        )
    buffer = BytesIO()
    workbook.save(buffer)

    started = time.monotonic()
    summary = pricelist.import_price_list(db_session, buffer.getvalue(), today=TODAY, dry_run=True)
    elapsed = time.monotonic() - started

    assert summary.yeni == 500
    assert elapsed < 10, f"500 satır {elapsed:.1f} sn sürdü (kriter: <10 sn)"


# --- API katmanı -------------------------------------------------------------


@pytest.fixture
def api(db_session: Session, store: Store) -> Iterator[TestClient]:
    """Test oturumuna bağlı API istemcisi."""

    async def session_override() -> Any:
        yield db_session

    app = create_app()
    app.dependency_overrides[deps.get_session] = session_override
    with TestClient(app, raise_server_exceptions=False) as client:
        yield client


def _headers(api: TestClient, brand: str = "alessi") -> dict[str, str]:
    response = api.post("/auth/dev-login", json={"email": "mert@mokkalabs.com", "brand": brand})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def test_export_endpoint_returns_xlsx_attachment(api: TestClient) -> None:
    """Uç xlsx döner ve dosya adı markayı taşır."""
    response = api.get("/alessi/price-list/export", headers=_headers(api))

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument"
    )
    assert "alessi" in response.headers["content-disposition"]
    assert load_workbook(BytesIO(response.content))[pricelist.SHEET_NAME]["A1"].value == (
        pricelist.TEMPLATE_VERSION
    )


def test_import_endpoint_roundtrip_reports_no_changes(api: TestClient) -> None:
    """Uçtan uca: export → import(dry_run) → 0/0/0 (spec §12A.6)."""
    exported = api.get("/alessi/price-list/export", headers=_headers(api)).content

    response = api.post(
        "/alessi/price-list/import",
        params={"dry_run": True},
        files={"file": ("fiyat.xlsx", exported)},
        headers=_headers(api),
    )

    assert response.status_code == 200, response.text
    payload = response.json()
    assert (payload["yeni"], payload["guncelleme"], payload["hata"]) == (0, 0, 0)
    assert payload["dry_run"] is True


def test_import_is_logged_as_batch(api: TestClient, db_session: Session) -> None:
    """§12A.2.4: her import `import_batches`'e loglanır (dry_run dahil)."""
    exported = api.get("/alessi/price-list/export", headers=_headers(api)).content

    api.post(
        "/alessi/price-list/import",
        params={"dry_run": True},
        files={"file": ("fiyat.xlsx", exported)},
        headers=_headers(api),
    )

    batch = db_session.scalar(select(ImportBatch).order_by(ImportBatch.created_at.desc()))
    assert batch is not None
    assert batch.kind == "price_list"
    assert batch.dry_run is True
    assert batch.filename == "fiyat.xlsx"


def test_import_of_foreign_template_returns_422(api: TestClient) -> None:
    """Şablon uyumsuzsa 422 + anlaşılır mesaj."""
    response = api.post(
        "/alessi/price-list/import",
        files={"file": ("baska.xlsx", b"excel degil")},
        headers=_headers(api),
    )

    assert response.status_code == 422
    assert "okunamadı" in response.json()["detail"]


def test_error_report_endpoint_returns_marked_file(api: TestClient) -> None:
    """§12A.2.3: hata raporu ucu "Hatalar" sayfalı dosyayı döndürür."""
    exported = api.get("/alessi/price-list/export", headers=_headers(api)).content
    broken = _edit(exported, pricelist.FIRST_DATA_ROW, "KDV %", 7)

    response = api.post(
        "/alessi/price-list/import/errors",
        files={"file": ("fiyat.xlsx", broken)},
        headers=_headers(api),
    )

    assert response.status_code == 200, response.text
    assert pricelist.ERROR_SHEET_NAME in load_workbook(BytesIO(response.content)).sheetnames
