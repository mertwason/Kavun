"""KVN-11: taslak ürün akışı (spec §12A.3).

Kabul kriteri (§12A.6): "taslak → promote → sipariş kârı hesaplama zinciri uçtan uca
testte kırılmaz" — dosyanın sonundaki zincir testi bunu doğrular.
"""

from __future__ import annotations

import uuid
from collections.abc import Iterator
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api import deps
from app.core.context import RequestContext, system_scope, use_context
from app.main import create_app
from app.models.catalog import Product, SkuCost, SkuLogistics, SkuPrice
from app.models.enums import DraftStatus, UserRole
from app.models.identity import Brand, Store
from app.models.results import LineProfit
from app.models.transactions import OrderLine
from app.models.workspace import ProductDraft
from app.services import drafts, pricelist
from app.services import profit as profit_service
from tests.profit_factories import ORDER_DATE, make_commission, make_order, make_store

D = Decimal
TODAY = date(2026, 8, 19)
CATEGORY = "Kahve/Harman"


@pytest.fixture
def store(db_session: Session) -> Iterator[Store]:
    """Mağaza + kategori tarifesi; marka bağlamı kurulur."""
    with system_scope():
        store = make_store(db_session)
        make_commission(db_session, store, rate=D("0.2000"), category=CATEGORY)
        brand = db_session.get(Brand, store.brand_id)
    assert brand is not None
    context = RequestContext(
        tenant_id=brand.tenant_id,
        user_id=None,
        brand_id=brand.id,
        brand_slug=brand.slug,
        role=UserRole.ADMIN,
    )
    with use_context(context):
        yield store


def _draft(db_session: Session, store: Store, **overrides: Any) -> ProductDraft:
    values: dict[str, Any] = {
        "name": "Yeni Harman 250g",
        "sku_onerisi": f"TASLAK-{uuid.uuid4().hex[:6].upper()}",
        "alis_maliyeti": D("60.0000"),
        "hedef_satis_fiyati": D("199.0000"),
        "kanal": "trendyol",
        "kategori": CATEGORY,
        "vat_rate": D("1.00"),
        "desi": D("1.00"),
    }
    values.update(overrides)
    return drafts.create_draft(
        db_session, tenant_id=store.tenant_id, brand_id=store.brand_id, **values
    )


# --- analiz (spec §12A.3) ----------------------------------------------------


def test_analysis_uses_category_tariff_for_commission(db_session: Session, store: Store) -> None:
    """§12A.3: komisyon kategori tarifesinden tahmin edilir (taslakta ürün henüz yok)."""
    analysis = drafts.analyze(
        db_session,
        price=D("199.00"),
        unit_cost=D("60.00"),
        vat_rate=D("1.00"),
        channel="trendyol",
        category=CATEGORY,
        cargo_cost=D("50.00"),
        on_date=TODAY,
    )

    assert analysis.commission_rate == D("0.2000")
    assert analysis.breakdown.cost_commission == D("39.8000")  # 199 × %20


def test_analysis_without_category_flags_missing_commission(
    db_session: Session, store: Store
) -> None:
    """Kategori yoksa uydurma oran kullanılmaz; satır uyarı taşır."""
    analysis = drafts.analyze(
        db_session,
        price=D("199.00"),
        unit_cost=D("60.00"),
        vat_rate=D("1.00"),
        channel="trendyol",
        category=None,
        cargo_cost=D("50.00"),
        on_date=TODAY,
    )

    assert analysis.commission_rate is None
    assert "komisyon_orani_yok" in analysis.breakdown.warnings
    assert analysis.breakdown.cost_commission == D("0")


def test_missing_cargo_estimate_is_flagged_not_invented(db_session: Session, store: Store) -> None:
    """Desi bazlı tarife KVN-14'te; şimdilik kargo verilmezse SIFIR + uyarı."""
    analysis = drafts.analyze(
        db_session,
        price=D("199.00"),
        unit_cost=D("60.00"),
        vat_rate=D("1.00"),
        channel="trendyol",
        category=CATEGORY,
        cargo_cost=None,
        on_date=TODAY,
    )

    assert analysis.cargo_cost == D("0")
    assert "kargo_tarifesi_yok" in analysis.breakdown.warnings


def test_analysis_uses_store_service_fee(db_session: Session, store: Store) -> None:
    """Hizmet bedeli mağaza ayarından gelir — form alanı değil."""
    store.service_fee_per_order = D("8.9900")
    db_session.flush()

    analysis = drafts.analyze(
        db_session,
        price=D("199.00"),
        unit_cost=D("60.00"),
        vat_rate=D("1.00"),
        channel="trendyol",
        category=CATEGORY,
        cargo_cost=D("0"),
        on_date=TODAY,
    )

    assert analysis.service_fee == D("8.9900")
    assert analysis.breakdown.cost_service_fee == D("8.9900")


def test_analysis_matches_engine_for_the_same_inputs(db_session: Session, store: Store) -> None:
    """Analiz motorun çıktısıdır — ikinci bir formül yok (CLAUDE.md §1)."""
    from app.engine.profit import LineInput, compute_line_profit

    analysis = drafts.analyze(
        db_session,
        price=D("199.00"),
        unit_cost=D("60.00"),
        vat_rate=D("1.00"),
        channel="trendyol",
        category=CATEGORY,
        cargo_cost=D("50.00"),
        on_date=TODAY,
    )
    expected = compute_line_profit(
        LineInput(
            line_gross=D("199.00"),
            qty=1,
            vat_percent=D("1.00"),
            unit_cost_net=D("60.00"),
            commission_rate=D("0.2000"),
            cargo_cost=D("50.00"),
            service_fee=store.service_fee_per_order or D("0"),
        )
    )

    assert analysis.breakdown.profit == expected.profit
    assert analysis.breakdown.margin_pct == expected.margin_pct


# --- promote (spec §12A.3) ---------------------------------------------------


def test_promote_creates_product_cost_logistics_and_price(
    db_session: Session, store: Store
) -> None:
    """§12A.3: promote → products + sku_costs + sku_logistics (+ fiyat) kayıtları."""
    draft = _draft(db_session, store)

    product = drafts.promote(db_session, draft, today=TODAY, user="mert@mokkalabs.com")

    assert draft.status is DraftStatus.PROMOTED
    assert draft.promoted_product_id == product.id
    assert product.sku == draft.sku_onerisi
    assert product.category == CATEGORY
    cost = db_session.scalar(select(SkuCost).where(SkuCost.product_id == product.id))
    assert cost is not None and cost.unit_cost == D("60.0000")
    assert (
        db_session.scalar(select(SkuLogistics).where(SkuLogistics.product_id == product.id))
        is not None
    )
    price = db_session.scalar(select(SkuPrice).where(SkuPrice.product_id == product.id))
    assert price is not None and price.price == D("199.0000")


def test_promote_without_sku_is_rejected(db_session: Session, store: Store) -> None:
    """SKU önerisi yoksa akış reddedilir — sessizce SKU uydurulmaz."""
    draft = _draft(db_session, store, sku_onerisi=None)

    with pytest.raises(drafts.DraftError, match="SKU önerisi boş"):
        drafts.promote(db_session, draft, today=TODAY)

    assert draft.status is DraftStatus.DRAFT


def test_promote_with_existing_sku_is_rejected(db_session: Session, store: Store) -> None:
    """Çakışan SKU mevcut ürünü EZMEZ; taslak taslak kalır."""
    first = _draft(db_session, store, sku_onerisi="CAKISAN-1")
    drafts.promote(db_session, first, today=TODAY)
    second = _draft(db_session, store, sku_onerisi="CAKISAN-1")

    with pytest.raises(drafts.DraftError, match="zaten kullanılıyor"):
        drafts.promote(db_session, second, today=TODAY)

    assert second.status is DraftStatus.DRAFT


def test_promote_twice_is_rejected(db_session: Session, store: Store) -> None:
    """Aynı taslak iki kez ürüne dönüşemez (çift ürün kaydı olmaz)."""
    draft = _draft(db_session, store)
    drafts.promote(db_session, draft, today=TODAY)

    with pytest.raises(drafts.DraftError, match="zaten"):
        drafts.promote(db_session, draft, today=TODAY)


def test_discard_keeps_the_record(db_session: Session, store: Store) -> None:
    """İptal kaydı silmez — geçmiş korunur (CLAUDE.md §1)."""
    draft = _draft(db_session, store)

    drafts.discard(db_session, draft)

    assert draft.status is DraftStatus.DISCARDED
    assert db_session.scalar(select(ProductDraft).where(ProductDraft.id == draft.id)) is not None


def test_promoted_draft_cannot_be_discarded(db_session: Session, store: Store) -> None:
    """Ürüne dönüşmüş taslak iptal edilemez."""
    draft = _draft(db_session, store)
    drafts.promote(db_session, draft, today=TODAY)

    with pytest.raises(drafts.DraftError, match="iptal edilemez"):
        drafts.discard(db_session, draft)


# --- Excel'den taslak (spec §12A.3) -----------------------------------------


def test_excel_rows_without_sku_become_drafts(db_session: Session, store: Store) -> None:
    """§12A.3: `as_draft=true` iken SKU'suz satırlar ürün değil TASLAK olur."""
    brand = db_session.get(Brand, store.brand_id)
    assert brand is not None
    payload = pricelist.export_price_list(db_session, brand_name=brand.name, today=TODAY)
    workbook = load_workbook(BytesIO(payload))
    sheet = workbook[pricelist.SHEET_NAME]
    sheet.append((None, "SKU'suz Aday Ürün", brand.name, "trendyol", 20, 1, 40, 129))
    buffer = BytesIO()
    workbook.save(buffer)

    summary = pricelist.import_price_list(
        db_session,
        buffer.getvalue(),
        today=TODAY,
        dry_run=False,
        as_draft=True,
        tenant_id=store.tenant_id,
        brand_id=store.brand_id,
    )

    assert summary.taslak == 1
    assert summary.hata == 0
    draft = db_session.scalar(select(ProductDraft).where(ProductDraft.name == "SKU'suz Aday Ürün"))
    assert draft is not None
    assert draft.hedef_satis_fiyati == D("129")
    assert db_session.scalar(select(Product).where(Product.sku == "")) is None


def test_excel_rows_without_sku_are_errors_when_not_draft_mode(
    db_session: Session, store: Store
) -> None:
    """`as_draft` kapalıyken SKU'suz satır hata olarak kalır (varsayılan davranış)."""
    brand = db_session.get(Brand, store.brand_id)
    assert brand is not None
    payload = pricelist.export_price_list(db_session, brand_name=brand.name, today=TODAY)
    workbook = load_workbook(BytesIO(payload))
    workbook[pricelist.SHEET_NAME].append((None, "Adsız", brand.name, "trendyol", 20, 1, 40, 129))
    buffer = BytesIO()
    workbook.save(buffer)

    summary = pricelist.import_price_list(db_session, buffer.getvalue(), today=TODAY)

    assert summary.hata == 1
    assert summary.taslak == 0


# --- kabul kriteri: uçtan uca zincir (spec §12A.6) --------------------------


def test_draft_to_promote_to_order_profit_chain(db_session: Session, store: Store) -> None:
    """§12A.6 kabul kriteri: taslak → promote → sipariş kârı zinciri kırılmaz.

    Taslağın analizindeki kâr ile, aynı fiyatla satılan siparişin motor kârı
    (kargo ve hizmet bedeli aynı verildiğinde) birebir aynı olmalıdır.
    """
    store.service_fee_per_order = D("0")
    db_session.flush()
    draft = _draft(db_session, store, sku_onerisi="ZINCIR-1")
    analysis = drafts.analyze_draft(db_session, draft, cargo_cost=D("0"), on_date=TODAY)

    product = drafts.promote(db_session, draft, today=ORDER_DATE.date(), user=None)
    order = make_order(db_session, store, [(product, 1, D("199.0000"))], cargo=None)
    profit_service.recompute_orders(db_session, order_ids=[order.id])

    order_line = db_session.scalar(select(OrderLine).where(OrderLine.order_id == order.id))
    assert order_line is not None
    record = db_session.scalar(select(LineProfit).where(LineProfit.order_line_id == order_line.id))
    assert record is not None
    assert record.profit == analysis.breakdown.profit
    assert record.cost_cogs == analysis.breakdown.cost_cogs


# --- API katmanı -------------------------------------------------------------


@pytest.fixture
def api(db_session: Session, store: Store) -> Iterator[TestClient]:
    """Test oturumuna bağlı API istemcisi."""

    async def session_override() -> Any:
        yield db_session

    app = create_app()
    app.dependency_overrides[deps.get_session] = session_override
    with TestClient(app, raise_server_exceptions=False) as client:
        yield client


def _headers(api: TestClient, brand: str = "alessi") -> dict[str, str]:
    response = api.post("/auth/dev-login", json={"email": "mert@mokkalabs.com", "brand": brand})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def _payload() -> dict[str, Any]:
    return {
        "name": "API Taslağı",
        "sku_onerisi": "API-TASLAK-1",
        "alis_maliyeti": "60.00",
        "hedef_satis_fiyati": "199.00",
        "kanal": "trendyol",
        "kategori": CATEGORY,
        "vat_rate": "1.00",
        "desi": "1.00",
        "kargo_tahmini": "50.00",
    }


def test_analyze_endpoint_writes_nothing(api: TestClient, db_session: Session) -> None:
    """§12A.5: form doldurulurken analiz alınır ama kayıt oluşmaz."""
    response = api.post("/alessi/drafts/analyze", json=_payload(), headers=_headers(api))

    assert response.status_code == 200, response.text
    assert Decimal(response.json()["profit"]) != 0
    assert db_session.scalar(select(ProductDraft)) is None


def test_create_and_promote_via_api(api: TestClient, db_session: Session) -> None:
    """Taslak kaydı → ürüne dönüştürme uçtan uca."""
    created = api.post("/alessi/drafts", json=_payload(), headers=_headers(api))
    assert created.status_code == 201, created.text
    draft_id = created.json()["id"]
    assert created.json()["analysis"]["commission_rate"] == "0.2000"

    promoted = api.post(f"/alessi/drafts/{draft_id}/promote", headers=_headers(api))

    assert promoted.status_code == 200, promoted.text
    assert promoted.json()["sku"] == "API-TASLAK-1"
    assert db_session.scalar(select(Product).where(Product.sku == "API-TASLAK-1")) is not None


def test_promote_conflict_returns_422(api: TestClient) -> None:
    """Çakışan SKU 422 + anlaşılır mesaj döner."""
    first = api.post("/alessi/drafts", json=_payload(), headers=_headers(api)).json()
    api.post(f"/alessi/drafts/{first['id']}/promote", headers=_headers(api))
    second = api.post("/alessi/drafts", json=_payload(), headers=_headers(api)).json()

    response = api.post(f"/alessi/drafts/{second['id']}/promote", headers=_headers(api))

    assert response.status_code == 422
    assert "zaten kullanılıyor" in response.json()["detail"]


def test_draft_of_other_brand_returns_404(api: TestClient, db_session: Session) -> None:
    """§3A.6: başka markanın taslağı 404 döner."""
    created = api.post("/alessi/drafts", json=_payload(), headers=_headers(api)).json()

    response = api.post(
        f"/kahveji/drafts/{created['id']}/promote", headers=_headers(api, "kahveji")
    )

    assert response.status_code == 404


def test_negative_price_is_rejected_by_validation(api: TestClient) -> None:
    """Negatif hedef fiyat şema seviyesinde reddedilir."""
    payload = _payload() | {"hedef_satis_fiyati": "-1"}

    response = api.post("/alessi/drafts", json=payload, headers=_headers(api))

    assert response.status_code == 422
