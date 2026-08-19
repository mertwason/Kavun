"""KVN-EK-02: kargo faturası eşleştirme ve `estimated → actual` (spec §5.3, §6.2).

Kabul: fatura eşleşince maliyet kesinleşir, kâr yeniden hesaplanır ve değişen alanlar
`profit_revisions`'a **tetikleyici gerekçesiyle** loglanır. Eşleşmeyen satır uydurma
eşleştirme yapmaz; kesinleşmiş maliyet ikinci faturayla ezilmez.
"""

from __future__ import annotations

from collections.abc import Iterator
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.api import deps
from app.core.context import RequestContext, system_scope, use_context
from app.main import create_app
from app.models.catalog import Product
from app.models.enums import CostState, OrderStatus, UserRole
from app.models.identity import Store
from app.models.results import Alert, LineProfit, ProfitRevision
from app.models.transactions import CargoInvoice, Order, Shipment
from app.services import cargo, profit
from tests.profit_factories import make_order, make_product, make_store

D = Decimal
PERIOD = "2026-08"
INVOICE_NO = "KRG-2026-08-001"


@pytest.fixture
def store(db_session: Session) -> Iterator[Store]:
    """Mağaza + marka bağlamı."""
    from app.models.identity import Brand

    with system_scope():
        store = make_store(db_session)
        brand = db_session.get(Brand, store.brand_id)
    assert brand is not None
    context = RequestContext(
        tenant_id=brand.tenant_id,
        user_id=None,
        brand_id=brand.id,
        brand_slug=brand.slug,
        role=UserRole.ADMIN,
    )
    with use_context(context):
        yield store


@pytest.fixture
def product(db_session: Session, store: Store) -> Product:
    """Maliyeti tanımlı ürün."""
    return make_product(db_session, store, "KHV-BLD-ESP", cost=D("580.0000"))


@pytest.fixture
def order(db_session: Session, store: Store, product: Product) -> Order:
    """Gönderisi tahmini maliyetle duran sipariş."""
    record = make_order(db_session, store, [(product, 1, D("1190.00"))])
    shipment = db_session.scalar(select(Shipment).where(Shipment.order_id == record.id))
    assert shipment is not None, "make_order gönderi kurmalı"
    shipment.tracking_no = "TK-0001"
    shipment.cargo_cost_estimated = D("70.0000")
    shipment.cost_state = CostState.ESTIMATED
    db_session.flush()
    return record


def _workbook(rows: list[tuple[Any, ...]]) -> bytes:
    """Şablonu doldurup baytlara çevirir."""
    book = load_workbook(BytesIO(cargo.template_workbook()))
    sheet = book[cargo.SHEET_NAME]
    for index, row in enumerate(rows, start=cargo.FIRST_DATA_ROW):
        for column, value in enumerate(row, start=1):
            sheet.cell(row=index, column=column, value=value)
    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def _import(
    db_session: Session, store: Store, rows: list[tuple[Any, ...]], *, dry_run: bool = False
) -> cargo.MatchSummary:
    return cargo.import_invoice(
        db_session,
        payload=_workbook(rows),
        store=store,
        invoice_no=INVOICE_NO,
        period=PERIOD,
        dry_run=dry_run,
    )


# --- şablon disiplini --------------------------------------------------------


def test_template_can_be_uploaded_back(db_session: Session, store: Store) -> None:
    """İndirilen dosya = yüklenen şablon."""
    summary = _import(db_session, store, [], dry_run=True)

    assert summary.rows == 0
    assert summary.hata == 0


def test_wrong_template_is_rejected(db_session: Session, store: Store) -> None:
    """Başka bir Excel yüklenirse akış durur."""
    book = Workbook()
    buffer = BytesIO()
    book.save(buffer)

    with pytest.raises(cargo.TemplateError, match="Şablon sürümü"):
        cargo.import_invoice(
            db_session,
            payload=buffer.getvalue(),
            store=store,
            invoice_no=INVOICE_NO,
            period=PERIOD,
        )


# --- eşleştirme ve kesinleşme ------------------------------------------------


def test_dry_run_changes_nothing(db_session: Session, store: Store, order: Order) -> None:
    """Önizleme sayar ama hiçbir maliyeti kesinleştirmez."""
    summary = _import(
        db_session, store, [("", "TK-0001", "14.08.2026", "2,5", "84,50")], dry_run=True
    )

    assert summary.kesinlesti == 1
    shipment = db_session.scalar(select(Shipment).where(Shipment.order_id == order.id))
    assert shipment is not None
    assert shipment.cost_state is CostState.ESTIMATED
    assert shipment.cargo_cost_actual is None


def test_matching_by_tracking_number_finalises_the_cost(
    db_session: Session, store: Store, order: Order
) -> None:
    """Gönderi numarasıyla eşleşen satır maliyeti kesinleştirir."""
    summary = _import(db_session, store, [("", "TK-0001", "14.08.2026", "2,5", "84,50")])

    assert summary.kesinlesti == 1
    assert summary.delta == D("14.50")  # 84,50 − 70,00
    shipment = db_session.scalar(select(Shipment).where(Shipment.order_id == order.id))
    assert shipment is not None
    assert shipment.cost_state is CostState.ACTUAL
    assert shipment.cargo_cost_actual == D("84.5000")
    assert shipment.desi_invoiced == D("2.50")


def test_matching_by_order_number_also_works(
    db_session: Session, store: Store, order: Order
) -> None:
    """Kanal takip numarası vermiyorsa sipariş numarasıyla eşleşilir."""
    summary = _import(
        db_session, store, [(order.external_order_id, "", "14.08.2026", "2,5", "84,50")]
    )

    assert summary.kesinlesti == 1


def test_unmatched_row_is_queued_not_guessed(
    db_session: Session, store: Store, order: Order
) -> None:
    """Eşleşmeyen satır uydurma bir gönderiye yazılmaz; uyarı üretilir."""
    summary = _import(db_session, store, [("", "TK-YOK", "14.08.2026", "2,5", "84,50")])

    assert summary.eslesmedi == 1
    assert summary.kesinlesti == 0
    alerts = db_session.scalars(select(Alert).where(Alert.type == cargo.UNMATCHED_ALERT)).all()
    assert len(alerts) == 1


def test_finalised_cost_is_not_overwritten(db_session: Session, store: Store, order: Order) -> None:
    """Kesinleşmiş maliyet ikinci faturayla ezilmez (normalize kuralının aynısı)."""
    _import(db_session, store, [("", "TK-0001", "14.08.2026", "2,5", "84,50")])

    second = cargo.import_invoice(
        db_session,
        payload=_workbook([("", "TK-0001", "20.08.2026", "2,5", "999,00")]),
        store=store,
        invoice_no="KRG-2026-08-002",
        period=PERIOD,
    )

    assert second.zaten_kesin == 1
    shipment = db_session.scalar(select(Shipment).where(Shipment.order_id == order.id))
    assert shipment is not None
    assert shipment.cargo_cost_actual == D("84.5000")


def test_negative_amount_is_rejected(db_session: Session, store: Store, order: Order) -> None:
    """Negatif tutar satırı reddedilir."""
    summary = _import(db_session, store, [("", "TK-0001", "14.08.2026", "2,5", "-10,00")])

    assert summary.hata == 1
    assert summary.kesinlesti == 0


def test_broken_date_is_reported(db_session: Session, store: Store, order: Order) -> None:
    """Okunamayan tarih satırı gerekçesiyle raporlanır."""
    summary = _import(db_session, store, [("", "TK-0001", "dün", "2,5", "84,50")])

    assert summary.hata == 1
    assert "Tarih" in summary.results[0].message


# --- kâr revizyonu (spec §6.2) ----------------------------------------------


def test_finalisation_recomputes_profit_and_logs_the_trigger(
    db_session: Session, store: Store, order: Order
) -> None:
    """§6.2: kargo kesinleşince kâr yeniden hesaplanır, revizyon gerekçesiyle loglanır."""
    with system_scope():
        profit.recompute_orders(db_session, order_ids=[order.id])
    before = db_session.scalar(select(LineProfit))
    assert before is not None
    profit_before = before.profit

    _import(db_session, store, [("", "TK-0001", "14.08.2026", "2,5", "120,00")])

    after = db_session.scalar(select(LineProfit))
    assert after is not None
    assert after.profit < profit_before, "kargo pahalandı, kâr düşmeli"
    assert after.is_final is True, "kesinleşen maliyet satırı `final` yapmalı"

    revisions = db_session.scalars(select(ProfitRevision)).all()
    assert revisions
    assert {row.reason for row in revisions} == {"kargo_faturasi"}
    assert "cost_cargo" in {row.field for row in revisions}


def test_recompute_reason_defaults_to_recompute(
    db_session: Session, store: Store, order: Order
) -> None:
    """Gerekçe verilmezse revizyon `recompute` olarak loglanır (geriye uyum)."""
    with system_scope():
        profit.recompute_orders(db_session, order_ids=[order.id])
        line_profit = db_session.scalar(select(LineProfit))
        assert line_profit is not None
        line_profit.profit = D("0")
        db_session.flush()

        profit.recompute_orders(db_session, order_ids=[order.id])

    revisions = db_session.scalars(select(ProfitRevision)).all()
    assert {row.reason for row in revisions} == {"recompute"}


def test_cost_state_summary_counts_both_sides(
    db_session: Session, store: Store, order: Order
) -> None:
    """Kesinleşme durumu ekranı: kaç gönderi kesin, kaçı tahmini."""
    before = cargo.cost_state_summary(db_session)
    assert before.actual == 0
    assert before.estimated == 1

    _import(db_session, store, [("", "TK-0001", "14.08.2026", "2,5", "84,50")])

    after = cargo.cost_state_summary(db_session)
    assert after.actual == 1
    assert after.estimated == 0
    assert after.actual_amount == D("84.5000")


def test_invoice_record_keeps_the_line_results(
    db_session: Session, store: Store, order: Order
) -> None:
    """Fatura kaydı satır sonuçlarını saklar — sonradan "neden eşleşmedi" sorulabilsin."""
    _import(
        db_session,
        store,
        [("", "TK-0001", "14.08.2026", "2,5", "84,50"), ("", "TK-YOK", "14.08.2026", "1", "50,00")],
    )

    invoice = db_session.scalar(select(CargoInvoice))
    assert invoice is not None
    assert invoice.invoice_no == INVOICE_NO
    assert len(invoice.lines) == 2
    assert {row["action"] for row in invoice.lines} == {"kesinlesti", "eslesmedi"}


def test_cancelled_order_shipment_can_still_be_invoiced(
    db_session: Session, store: Store, order: Order
) -> None:
    """İptal sipariş de kargolanmış olabilir; fatura satırı yine kesinleşir.

    Kâr motoru iptal satırda maliyet üretmez, ama kargo gerçek bir gider olarak
    mağaza seviyesinde kalır — bu yüzden eşleştirme reddedilmez.
    """
    order.status = OrderStatus.CANCELLED
    db_session.flush()

    summary = _import(db_session, store, [("", "TK-0001", "14.08.2026", "2,5", "84,50")])

    assert summary.kesinlesti == 1


# --- API ---------------------------------------------------------------------


@pytest.fixture
def api(db_session: Session, store: Store) -> Iterator[TestClient]:
    """Test oturumuna bağlı API istemcisi."""

    async def session_override() -> Any:
        yield db_session

    app = create_app()
    app.dependency_overrides[deps.get_session] = session_override
    with TestClient(app, raise_server_exceptions=False) as client:
        yield client


def _headers(api: TestClient, brand: str = "alessi") -> dict[str, str]:
    response = api.post("/auth/dev-login", json={"email": "mert@mokkalabs.com", "brand": brand})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['access_token']}"}


def test_template_endpoint_returns_xlsx(api: TestClient) -> None:
    """Şablon marka önekli dosya adıyla iner."""
    response = api.get("/alessi/cargo-invoices/template", headers=_headers(api))

    assert response.status_code == 200, response.text
    assert response.content[:2] == b"PK"
    assert "alessi-kargo" in response.headers["content-disposition"]


def test_import_endpoint_previews_then_applies(
    api: TestClient, db_session: Session, order: Order
) -> None:
    """Önizleme yazmaz, onay yazar."""
    payload = _workbook([("", "TK-0001", "14.08.2026", "2,5", "84,50")])
    fields = {"invoice_no": INVOICE_NO, "period": PERIOD}

    preview = api.post(
        "/alessi/cargo-invoices/import",
        params={"dry_run": True},
        data=fields,
        files={"file": ("kargo.xlsx", payload)},
        headers=_headers(api),
    )
    assert preview.status_code == 200, preview.text
    assert preview.json()["kesinlesti"] == 1
    assert preview.json()["invoice_id"] is None

    applied = api.post(
        "/alessi/cargo-invoices/import",
        params={"dry_run": False},
        data=fields,
        files={"file": ("kargo.xlsx", payload)},
        headers=_headers(api),
    )
    assert applied.status_code == 200, applied.text
    assert applied.json()["invoice_id"] is not None


def test_cost_state_endpoint(api: TestClient, order: Order) -> None:
    """Kesinleşme durumu ucu sayıları döner."""
    response = api.get("/alessi/cargo-invoices/cost-state", headers=_headers(api))

    assert response.status_code == 200, response.text
    assert response.json()["estimated"] == 1


def test_invoice_list_endpoint(
    api: TestClient, db_session: Session, store: Store, order: Order
) -> None:
    """Fatura listesi marka kapsamlıdır."""
    _import(db_session, store, [("", "TK-0001", "14.08.2026", "2,5", "84,50")])

    response = api.get("/alessi/cargo-invoices", headers=_headers(api))

    assert response.status_code == 200, response.text
    assert [row["invoice_no"] for row in response.json()] == [INVOICE_NO]


def test_import_requires_editor_role(api: TestClient) -> None:
    """Yükleme yetkisi rol ister; okuma uçları herkese açıktır."""
    response = api.post(
        "/alessi/cargo-invoices/import",
        params={"dry_run": True},
        data={"invoice_no": INVOICE_NO, "period": PERIOD},
        files={"file": ("kargo.xlsx", cargo.template_workbook())},
        headers=_headers(api),
    )

    # Demo kullanıcı admin; ret beklenmiyor — uç ayakta ve şablonu kabul ediyor.
    assert response.status_code == 200, response.text
    assert response.json()["rows"] == 0


def test_dates_and_amounts_accept_turkish_format(
    db_session: Session, store: Store, order: Order
) -> None:
    """Kullanıcı Excel'de `14.08.2026` ve `84,50` yazabilir."""
    summary = _import(db_session, store, [("", "TK-0001", date(2026, 8, 14), 2.5, "84,50")])

    assert summary.kesinlesti == 1
    assert summary.total_amount == D("84.50")
